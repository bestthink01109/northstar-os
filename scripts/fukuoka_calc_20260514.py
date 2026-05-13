import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings
import re
import os
import glob
from datetime import datetime
from get_attendance_data import fetch_attendance_data, sync_staff_list

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ==========================================
# ファイル選択ロジック
# ==========================================
# 「【出勤簿】」で始まり「_計算式適用済」を含まないExcelファイルを検索
xlsx_files = sorted(glob.glob("【出勤簿】*.xlsx"))
xlsx_files = [f for f in xlsx_files if "_計算式適用済" not in f]

if not xlsx_files:
    print("エラー: 処理対象の『【出勤簿】YYYYMM.xlsx』ファイルがこのフォルダに見つかりません。")
    exit()

if len(xlsx_files) == 1:
    file_path = xlsx_files[0]
    print(f">>> ファイル '{file_path}' を自動選択しました。")
else:
    print("\n" + "="*50)
    print(" 処理する出勤簿ファイルを選択してください")
    print("="*50)
    for i, f in enumerate(xlsx_files):
        print(f" [{i+1}] {f}")
    print("="*50)
    
    try:
        choice = int(input(f"番号を入力してください (1-{len(xlsx_files)}): "))
        file_path = xlsx_files[choice-1]
    except (ValueError, IndexError):
        print("!! 無効な選択です。プログラムを終了します。")
        exit()

# ファイル名から年月(YYYYMM)を抽出
match = re.search(r'(\d{4})(\d{2})', file_path)
if match:
    target_month = f"{match.group(1)}-{match.group(2)}"
else:
    target_month = datetime.now().strftime("%Y-%m")

print(f">>> {target_month} の実績データをLINEから取得して処理を開始します...")

# 成果物を同じフォルダ内に保存する設定に変更
output_path = "./" + file_path.replace(".xlsx", "_計算式適用済.xlsx")

try:
    wb = openpyxl.load_workbook(file_path)
except Exception as e:
    print(f"エラー: '{file_path}' を開けませんでした。({e})")
    exit()

exclude_sheets = ['設定', '集計', 'H', 'I', '実績']
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
auto_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
thin_side = Side(style='thin', color='000000')
medium_side = Side(style='medium', color='000000')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ws_settings = wb['設定']
employee_names = []
for row in range(5, 50):
    val = ws_settings.cell(row=row, column=2).value
    if val:
        employee_names.append(str(val))

# Googleスプレッドシート（社員マスター）へ同期
sync_staff_list(employee_names)

# ==========================================
# 1. マスター入力用「実績」シートの構築
# ==========================================
if '実績' in wb.sheetnames:
    wb.remove(wb['実績'])
ws_jisseki = wb.create_sheet(title='実績', index=0)

ws_jisseki['A1'] = "出勤簿 一括入力マスター"
ws_jisseki['A1'].font = Font(bold=True, size=14)
ws_jisseki['A2'] = "氏名"
ws_jisseki['B2'] = "項目"

for d in range(1, 32):
    c_let = openpyxl.utils.get_column_letter(d + 2)
    ws_jisseki[f'{c_let}2'] = d
    ws_jisseki[f'{c_let}2'].alignment = Alignment(horizontal="center")

ws_jisseki['AH2'] = "前月最終週\n不足枠"
ws_jisseki['AI2'] = "当月最終週\n不足枠"
ws_jisseki['AH2'].alignment = Alignment(wrap_text=True, horizontal="center")
ws_jisseki['AI2'].alignment = Alignment(wrap_text=True, horizontal="center")
ws_jisseki.column_dimensions['AH'].width = 12
ws_jisseki.column_dimensions['AI'].width = 12
ws_jisseki.column_dimensions['A'].width = 12
ws_jisseki.column_dimensions['B'].width = 15

# 【改善】状態のリストをシンプルにし、場所の指定は2行目(フリー入力)で行う
dv_shukkin = DataValidation(type="list", formula1='"現場,出張,休出,休出(出張),振出,振出(出張),振休,休日,有給,欠勤,特別"', allow_blank=True)
ws_jisseki.add_data_validation(dv_shukkin)
dv_bento = DataValidation(type="list", formula1='"〇"', allow_blank=True)
ws_jisseki.add_data_validation(dv_bento)

# LINE連携システム(GAS)から実績データを取得
inject_data = fetch_attendance_data(target_month)

current_row = 3
emp_rows = {}

for name in employee_names:
    ws_jisseki.merge_cells(f'A{current_row}:A{current_row+3}')
    ws_jisseki[f'A{current_row}'] = name
    ws_jisseki[f'A{current_row}'].alignment = Alignment(vertical="center", horizontal="center")
    
    ws_jisseki[f'B{current_row}'] = "出勤状態"
    ws_jisseki[f'B{current_row+1}'] = "現場名"
    ws_jisseki[f'B{current_row+2}'] = "残業"
    ws_jisseki[f'B{current_row+3}'] = "弁当"
    
    clean_n = name.replace(" ", "").replace("　", "")
    emp_rows[clean_n] = current_row
    person_data = inject_data.get(clean_n, {})
    
    for d in range(1, 32):
        c_let = openpyxl.utils.get_column_letter(d + 2)
        status = ""
        location = ""
        # inject_data内の全てのキーをループして、該当日(d)が含まれているかチェック
        for key, days in person_data.items():
            if key in ["残業", "弁当"]: continue
            if d in days:
                if "(" in key and ")" in key:
                    status = key.split("(")[0]
                    location = key.split("(")[1].replace(")", "")
                else:
                    status = key
                    location = ""
                break
        
        if status: ws_jisseki[f'{c_let}{current_row}'] = status
        if location: ws_jisseki[f'{c_let}{current_row+1}'] = location
        if d in person_data.get("残業", {}): ws_jisseki[f'{c_let}{current_row+2}'] = person_data["残業"][d]
        if d in person_data.get("弁当", []): ws_jisseki[f'{c_let}{current_row+3}'] = "〇"
            
        for offset in range(4):
            ws_jisseki[f'{c_let}{current_row+offset}'].fill = input_fill
            ws_jisseki[f'{c_let}{current_row+offset}'].border = thin_border
            
    # XML破損対策：DataValidationを1セルずつではなく、行範囲（C列〜AG列）で一括適用する
    dv_shukkin.add(f'C{current_row}:AG{current_row}')
    dv_bento.add(f'C{current_row+3}:AG{current_row+3}')
    
    
    ws_jisseki[f'AH{current_row}'].fill = input_fill
    ws_jisseki[f'AH{current_row}'].border = thin_border
    ws_jisseki[f'AH{current_row}'].number_format = '[h]:mm'
    
    ws_jisseki[f'AI{current_row}'].fill = auto_fill
    ws_jisseki[f'AI{current_row}'].border = thin_border
    ws_jisseki[f'AI{current_row}'].number_format = '[h]:mm'
    ws_jisseki[f'AI{current_row}'] = f"='{name}'!$AJ$1"
            
    current_row += 4

ws_jisseki.freeze_panes = 'C3'

# ==========================================
# 2. 各シートへの流し込み（安全装置版）
# ==========================================
template_sheet_name = next((s for s in wb.sheetnames if s not in exclude_sheets), None)
if template_sheet_name:
    template_ws = wb[template_sheet_name]
    for name in employee_names:
        if name not in wb.sheetnames:
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = name

# 計算式の定義
formulas = {
    'BA': '=IF($B{r}="","", $B{r} - WEEKDAY($B{r}, 1) + 1)',
    'BB': '=IF(AND($J{r}="", $L{r}=""), "", IF($J{r}<>"", $J{r}, $L{r}))',
    'BC': '=IF(AND($P{r}="", $N{r}=""), "", IF($P{r}<>"", $P{r}, $N{r}) + IF(IF($P{r}<>"", $P{r}, $N{r}) < IF($J{r}<>"", $J{r}, $L{r}), 1, 0))',
    'BD': '=IF(OR($D{r}="有給",$D{r}="欠勤",$D{r}="特別",$D{r}="休出",$D{r}="振休",AND(WEEKDAY($B{r},2)=7, LEFT($D{r},2)<>"振出"),BB{r}="",BC{r}=""), 0, MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))))',
    'BE': '=IF(BD{r}=0, 0, IF(OR(WEEKDAY($B{r},2)<=5, LEFT($D{r},2)="振出"), MIN(BD{r}, 7/24), MIN(BD{r}, 5/24)))',
    'BF': '=MAX(0, BD{r} - BE{r})',
    'BG': '=IF(OR(WEEKDAY($B{r},2)<=5, LEFT($D{r},2)="振出"), MIN(BF{r}, 1/24), MIN(BF{r}, 3/24))',
    'BH': '=IF($BA{r}="","", SUMIFS($BE$4:$BE$34, $BA$4:$BA$34, $BA{r}, $B$4:$B$34, "<"&$B{r}) + IF($BA{r}=$BA$4, IF(ISNUMBER(実績!$AH${base_row}), 実績!$AH${base_row}, 0), 0))',
    'BI': '=IF($BA{r}="","", IF(AND($BA{r}=MAX($BA$4:$BA$34), WEEKDAY(MAX($B$4:$B$34), 1)<7), SUMIFS($BL$4:$BL$34, $BA$4:$BA$34, $BA{r}), MAX(0, 40/24 - (SUMIFS($BE$4:$BE$34, $BA$4:$BA$34, $BA{r}) + IF($BA{r}=$BA$4, IF(ISNUMBER(実績!$AH${base_row}), 実績!$AH${base_row}, 0), 0)))))',
    'BJ': '=IF($BA{r}="","", SUMIFS($BG$4:$BG$34, $BA$4:$BA$34, $BA{r}, $B$4:$B$34, "<"&$B{r}))',
    'BK': '=IF($B{r}="","", IF(OR($D{r}="休出", AND(WEEKDAY($B{r},2)=7, LEFT($D{r},2)<>"振出"), $D{r}="振休"), 0, IF(OR(WEEKDAY($B{r},2)<=5, LEFT($D{r},2)="振出"), 7/24, 5/24)))',
    'BL': '=IF($B{r}="","", MAX(0, BK{r} - BE{r}))',
    'BM': '=IF(OR(AND($D{r}="出張", $F{r}<>""), AND($AA{r}="出張", $F{r}<>"")), IF(COUNTIFS($D$4:$D{r}, "出張", $F$4:$F{r}, $F{r}) + COUNTIFS($AA$4:$AA{r}, "出張", $F$4:$F{r}, $F{r})=1, ROW(), ""), "")',
    'R': '=IF($B{r}="","", IF(ROUND(MAX(0, MIN(BE{r}, 40/24 - BH{r})), 5)<=0, "", MAX(0, MIN(BE{r}, 40/24 - BH{r}))))',
    'S': '=IF($B{r}="","", IF(ROUND(MAX(0, MIN(BG{r}, BI{r} - BJ{r})), 5)<=0, "", MAX(0, MIN(BG{r}, BI{r} - BJ{r}))))',
    'T': '=IF($B{r}="","", IF(ROUND(BD{r} - N(R{r}) - N(S{r}), 5)<=0, "", BD{r} - N(R{r}) - N(S{r})))',
    'U': '=IF(OR($BB{r}="", $BC{r}=""), "", IF(ROUND(MAX(0, MIN($BC{r}, 1+5/24) - MAX($BB{r}, 22/24)) + MAX(0, MIN($BC{r}, 5/24) - $BB{r}), 5)<=0, "", MAX(0, MIN($BC{r}, 1+5/24) - MAX($BB{r}, 22/24)) + MAX(0, MIN($BC{r}, 5/24) - $BB{r})))',
    'V': '=IF($B{r}="","", IF(OR($D{r}="休出", AND(WEEKDAY($B{r},2)=7, LEFT($D{r},2)<>"振出")), IF(OR(BB{r}="", BC{r}=""), "", IF(ROUND(MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))), 5)<=0, "", MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))))), ""))',
    'W': '=IF($D{r}="有給", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")',
    'X': '=IF($B{r}="","", IF(OR($D{r}="有給",$D{r}="欠勤",$D{r}="特別",$D{r}="休出",$D{r}="振休",AND(WEEKDAY($B{r},2)=7, LEFT($D{r},2)<>"振出"),AND(BB{r}="",BC{r}="")), "", IF(OR(WEEKDAY($B{r},2)<=5, LEFT($D{r},2)="振出"), IF(ROUND(7/24-BE{r}, 5)>0, 7/24-BE{r}, ""), IF(ROUND(5/24-BE{r}, 5)>0, 5/24-BE{r}, ""))))',
    'Y': '=IF($D{r}="欠勤", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")',
    'Z': '=IF($D{r}="特別", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")'
}

for sheet_name in wb.sheetnames:
    if sheet_name in exclude_sheets: continue
    ws = wb[sheet_name]
    clean_sn = sheet_name.replace(" ", "").replace("　", "")
    base_row = emp_rows.get(clean_sn)
    if not base_row: continue

    # 作業列の定義（画像のあるAA列を避け、遠くのBZ列を使用）
    ws.column_dimensions['BZ'].width = 1
    # ws['BZ3'] = "作業列" # ヘッダー不要なら書かない
    
    # 列のクリアを止め、必要なセルだけを更新するように変更
    # for col in ['AG', 'AH', 'AI', 'AJ', 'AK', 'AL']: ... 消去

    ws['AH1'] = '="前月最終週の累計:"'
    ws['AH1'].alignment = Alignment(horizontal="right")
    ws['AJ2'] = f"=実績!$AH${base_row}"
    ws['AJ2'].number_format = '[h]:mm'
    ws['AJ2'].fill = input_fill
    ws['AJ2'].border = thin_border
    
    # AH2〜AH34 に数式を入れる処理は省略し、AH列はマスターからの転記のみとする
    # テンプレートにあるかもしれない数式をクリア
    for r in range(2, 35):
        ws[f'AH{r}'] = ""

    ws['AI1'] = '="当月最終週の累計:"'
    ws['AI1'].alignment = Alignment(horizontal="right")
    ws['AJ1'] = '=IF($B$4="","", IF(WEEKDAY(MAX($B$4:$B$34),1)=7, 0, SUMIFS($R$4:$R$34, $BA$4:$BA$34, MAX($BA$4:$BA$34)) + SUMIFS($S$4:$S$34, $BA$4:$BA$34, MAX($BA$4:$BA$34))))'
    ws['AJ1'].number_format = '[h]:mm'
    ws['AJ1'].fill = auto_fill
    ws['AJ1'].border = thin_border

    ws.column_dimensions['AK'].width = 8
    ws.column_dimensions['AL'].width = 8
    ws['AK3'] = '="例外出社"'
    ws['AK3'].font = Font(bold=True)
    ws['AK3'].alignment = Alignment(horizontal="center")
    ws['AK3'].fill = input_fill
    ws['AK3'].border = thin_border
    
    ws['AL3'] = '="例外退社"'
    ws['AL3'].font = Font(bold=True)
    ws['AL3'].alignment = Alignment(horizontal="center")
    ws['AL3'].fill = input_fill
    ws['AL3'].border = thin_border
    
    for row in range(4, 35):
        d_num = row - 3
        c_let = openpyxl.utils.get_column_letter(d_num + 2)
        
        # D列（状態）
        ws[f'D{row}'] = f'=IF(実績!{c_let}{base_row}="","",IF(実績!{c_let}{base_row}="休出(出張)","休出",IF(実績!{c_let}{base_row}="振出(出張)","振出",実績!{c_let}{base_row})))'
        # F列（現場名）
        ws[f'F{row}'] = f'=IF(実績!{c_let}{base_row+1}="","",実績!{c_let}{base_row+1})' 
        
        # 【BZ列（旧AA列）】内部計算用。画像保護のため遠くへ移動。
        ws[f'BZ{row}'] = f'=IF(OR(実績!{c_let}{base_row}="休出(出張)",実績!{c_let}{base_row}="振出(出張)"),"出張","")'
        
        ws[f'AJ{row}'] = f'=実績!{c_let}{base_row+3}'
        # 文字色を白にして見えなくする設定（必要なら維持、不要なら削除）
        # ws[f'AJ{row}'].font = Font(color="FFFFFF")
        
        for col, f_template in formulas.items():
            ws[f'{col}{row}'].value = f_template.format(r=row, base_row=base_row)
            
        ws[f'AK{row}'].fill = input_fill
        ws[f'AK{row}'].border = thin_border
        ws[f'AK{row}'].number_format = '[h]:mm'
        
        ws[f'AL{row}'].fill = input_fill
        ws[f'AL{row}'].border = thin_border
        ws[f'AL{row}'].number_format = '[h]:mm'

        if type(ws[f'J{row}']).__name__ != 'MergedCell':
            ws[f'J{row}'].value = f'=IF($AK{row}<>"",$AK{row},IF(OR($D{row}="現場",$D{row}="出張",$D{row}="休出",$D{row}="振出"),8/24,""))'
            ws[f'J{row}'].number_format = '[h]:mm'
        if type(ws[f'P{row}']).__name__ != 'MergedCell':
            ws[f'P{row}'].value = f'=IF($AL{row}<>"",$AL{row},IF(OR($D{row}="現場",$D{row}="出張",$D{row}="休出",$D{row}="振出"),IF(WEEKDAY($B{row},2)=6,15/24,17/24)+IF(ISNUMBER(実績!{c_let}{base_row+2}),実績!{c_let}{base_row+2}/24,0),IF(IF($J{row}<>"",$J{row},$L{row})<>"",IF($J{row}<>"",$J{row},$L{row})+IF(WEEKDAY($B{row},2)=6,7/24,9/24)+IF(ISNUMBER(実績!{c_let}{base_row+2}),実績!{c_let}{base_row+2}/24,0),"")))'
            ws[f'P{row}'].number_format = '[h]:mm'

    # C38: 出勤数の集計（現場、出張、休出に加えて「振出」もカウント）
    # フォントやデザインには触れず、数式だけを上書きする
    ws['C38'] = '=COUNTIFS($D$4:$D$34,"現場")+COUNTIFS($D$4:$D$34,"出張")+COUNTIFS($D$4:$D$34,"振出")'

    pass

wb.save("temp_output.xlsx")

# =====================================================================
# 【画像保護の禁じ手】
# openpyxlで消えてしまう画像をXMLレベルで強制復元する
# =====================================================================
try:
    from regex_patcher import surgical_regex_patch
    surgical_regex_patch(file_path, "temp_output.xlsx", output_path)
    if os.path.exists("temp_output.xlsx"):
        os.remove("temp_output.xlsx")
except Exception as e:
    print(f"超精密パッチでエラーが発生しました: {e}")
    # 失敗時は生成ファイルをそのまま成果物とする
    if os.path.exists("temp_output.xlsx"):
        os.rename("temp_output.xlsx", output_path)

print(f"完了しました！ '{output_path}' を確認してください。")

# =====================================================================
# 【openpyxl既知バグ 後処理修正】
# openpyxlはコピーシートで <mergeCells count> 属性を正しく更新しない。
# 保存後にZIPを直接操作し、実際の <mergeCell> 要素数と count を一致させる。
# これをしないと Excel が「一部の内容に問題が見つかりました」エラーを出す。
# =====================================================================
import zipfile, re, shutil, io

def fix_merge_counts(path):
    tmp_path = path + ".tmp.xlsx"
    with zipfile.ZipFile(path, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                    text = data.decode('utf-8')

                    # 修正1: <mergeCells count> を実要素数に合わせる
                    actual_count = len(re.findall(r'<mergeCell\b', text))
                    if actual_count > 0:
                        text = re.sub(
                            r'<mergeCells count="\d+"',
                            f'<mergeCells count="{actual_count}"',
                            text
                        )

                    # 修正2: 空の <dataValidations count="0"/> を除去
                    # openpyxlがDV削除後も空タグを残してしまう既知バグ。
                    # Mac版ExcelはこれをHRESULT 0x8000ffff (XMLエラー)として報告する。
                    text = re.sub(r'<dataValidations count="0"\s*/>', '', text)

                    data = text.encode('utf-8')
                zout.writestr(item, data)
    shutil.move(tmp_path, path)

fix_merge_counts(output_path)
print(f"完了しました！ '{output_path}' を確認してください。（XMLマージカウント修正済み）")