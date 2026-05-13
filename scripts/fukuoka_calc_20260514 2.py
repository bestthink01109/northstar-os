import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

file_path = "【出勤簿】202604.xlsx"
output_path = "【出勤簿】202604_計算式適用済.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    print(f"エラー: '{file_path}' が見つかりません。")
    exit()

exclude_sheets = ['設定', '集計', 'H', 'I', '実績']
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
auto_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
thin_side = Side(style='thin', color='000000')
medium_side = Side(style='medium', color='000000')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ws_settings = wb['設定']
employee_names = []
for row in range(5, 50):
    val = ws_settings.cell(row=row, column=2).value
    if val:
        employee_names.append(str(val))

# ==========================================
# 1. マスター入力用「実績」シートの構築
# ==========================================
if '実績' in wb.sheetnames:
    wb.remove(wb['実績'])
ws_jisseki = wb.create_sheet(title='実績', index=0)

ws_jisseki['A1'] = "出勤簿 一括入力マスター"
ws_jisseki['A1'].font = Font(bold=True, size=14)
ws_jisseki['A2'] = "氏名"
ws_jisseki['B2'] = "項目"

for d in range(1, 32):
    c_let = openpyxl.utils.get_column_letter(d + 2)
    ws_jisseki[f'{c_let}2'] = d
    ws_jisseki[f'{c_let}2'].alignment = Alignment(horizontal="center")

ws_jisseki['AH2'] = "前月最終週\n不足枠"
ws_jisseki['AI2'] = "当月最終週\n不足枠"
ws_jisseki['AH2'].alignment = Alignment(wrap_text=True, horizontal="center")
ws_jisseki['AI2'].alignment = Alignment(wrap_text=True, horizontal="center")
ws_jisseki.column_dimensions['AH'].width = 12
ws_jisseki.column_dimensions['AI'].width = 12
ws_jisseki.column_dimensions['A'].width = 12
ws_jisseki.column_dimensions['B'].width = 15

dv_shukkin = DataValidation(type="list", formula1='"現場(大牟田),現場(工場),出張(千葉),休出(出張),有給,欠勤,特別,休出"', allow_blank=True)
ws_jisseki.add_data_validation(dv_shukkin)
dv_bento = DataValidation(type="list", formula1='"〇"', allow_blank=True)
ws_jisseki.add_data_validation(dv_bento)

inject_data = {
    "土本凱斗": {"現場(大牟田)": [1,2,3,4,6,7,8,9,10,13,14,15,16,17,18,20,23,24,25], "現場(工場)": [21,22], "出張(千葉)": [27,28,29,30], "残業": {27: 2, 28: 2}, "弁当": [28]},
    "中嶋秀斗": {"現場(大牟田)": [1,2,3,4,6,7,8,9,10,13,14,15,16,17,18,20,21,22,23,24,25,27,28]},
    "坂井謙斗": {"現場(大牟田)": [1,2,3,4,6,7,8,9], "出張(千葉)": [12,13,14,15,16,17,18,20,21,22,23,24,25,27,28,29,30], "休出(出張)": [19], "残業": {13:2, 14:2, 16:2, 17:2, 20:2, 21:2, 22:2, 24:2, 27:2, 28:2}},
    "石村侑茉": {"現場(大牟田)": [1,2,3,4,6,7,8,9,10,13,14,15,16,17,18,20,21,22,23], "出張(千葉)": [27,28,29,30], "有給": [24,25], "残業": {28:2, 29:2, 30:2}, "弁当": [28,29,30]},
    "鳳ノ城翔音": {"現場(大牟田)": [1,2,3,6,7,8,10,13,14,15,16,17,18,20,21,22,23,24,25,27,28]},
    "内村巴瞬": {"現場(大牟田)": [1,2,3,6,7,8,9,10,13,14,15,16,17,18,20,21,22,23,24,25], "出張(千葉)": [27,28,29,30], "残業": {28: 2}, "弁当": [28,29]},
    "福山舞乙": {"現場(大牟田)": [1,2,3,4,6,7,8,9,10,13,14,15,16,17,18,20,21,22,23,24,25,27,28]}
}

current_row = 3
emp_rows = {}

for name in employee_names:
    ws_jisseki.merge_cells(f'A{current_row}:A{current_row+2}')
    ws_jisseki[f'A{current_row}'] = name
    ws_jisseki[f'A{current_row}'].alignment = Alignment(vertical="center", horizontal="center")
    
    ws_jisseki[f'B{current_row}'] = "出勤状態"
    ws_jisseki[f'B{current_row+1}'] = "残業(h)"
    ws_jisseki[f'B{current_row+2}'] = "弁当(〇)"
    
    clean_n = name.replace(" ", "").replace("　", "")
    emp_rows[clean_n] = current_row
    person_data = inject_data.get(clean_n, {})
    
    for d in range(1, 32):
        c_let = openpyxl.utils.get_column_letter(d + 2)
        status = ""
        if d in person_data.get("現場(大牟田)", []): status = "現場(大牟田)"
        elif d in person_data.get("現場(工場)", []): status = "現場(工場)"
        elif d in person_data.get("出張(千葉)", []): status = "出張(千葉)"
        elif d in person_data.get("休出(出張)", []): status = "休出(出張)"
        elif d in person_data.get("有給", []): status = "有給"
        
        if status: ws_jisseki[f'{c_let}{current_row}'] = status
        if d in person_data.get("残業", {}): ws_jisseki[f'{c_let}{current_row+1}'] = person_data["残業"][d]
        if d in person_data.get("弁当", []): ws_jisseki[f'{c_let}{current_row+2}'] = "〇"
            
        for offset in range(3):
            ws_jisseki[f'{c_let}{current_row+offset}'].fill = input_fill
            ws_jisseki[f'{c_let}{current_row+offset}'].border = thin_border
    
    ws_jisseki[f'AH{current_row}'].fill = input_fill
    ws_jisseki[f'AH{current_row}'].border = thin_border
    ws_jisseki[f'AH{current_row}'].number_format = '[h]:mm'
    
    ws_jisseki[f'AI{current_row}'].fill = auto_fill
    ws_jisseki[f'AI{current_row}'].border = thin_border
    ws_jisseki[f'AI{current_row}'].number_format = '[h]:mm'
    ws_jisseki[f'AI{current_row}'] = f"='{name}'!$AJ$1"
            
    current_row += 3

ws_jisseki.freeze_panes = 'C3'

# ==========================================
# 2. 各シートへの流し込み（安全装置版）
# ==========================================
template_sheet_name = next((s for s in wb.sheetnames if s not in exclude_sheets), None)
if template_sheet_name:
    template_ws = wb[template_sheet_name]
    for name in employee_names:
        if name not in wb.sheetnames:
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = name

# 計算式の定義
formulas = {
    'BA': '=IF($B{r}="","", $B{r} - WEEKDAY($B{r}, 2) + 1)',
    'BB': '=IF(AND($J{r}="", $L{r}=""), "", IF($J{r}<>"", $J{r}, $L{r}))',
    'BC': '=IF(AND($P{r}="", $N{r}=""), "", IF($P{r}<>"", $P{r}, $N{r}) + IF(IF($P{r}<>"", $P{r}, $N{r}) < IF($J{r}<>"", $J{r}, $L{r}), 1, 0))',
    'BD': '=IF(OR($D{r}="有給",$D{r}="欠勤",$D{r}="特別",$D{r}="休出",WEEKDAY($B{r},2)=7,BB{r}="",BC{r}=""), 0, MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))))',
    'BE': '=IF(BD{r}=0, 0, IF(WEEKDAY($B{r},2)<=5, MIN(BD{r}, 7/24), MIN(BD{r}, 5/24)))',
    'BF': '=MAX(0, BD{r} - BE{r})',
    'BG': '=IF(WEEKDAY($B{r},2)<=5, MIN(BF{r}, 1/24), MIN(BF{r}, 3/24))',
    'BH': '=IF($BA{r}="","", SUMIFS($W$4:$W$34, $BA$4:$BA$34, $BA{r}) + SUMIFS($X$4:$X$34, $BA$4:$BA$34, $BA{r}) + SUMIFS($Y$4:$Y$34, $BA$4:$BA$34, $BA{r}) + SUMIFS($Z$4:$Z$34, $BA$4:$BA$34, $BA{r}) + IF($BA{r}=$BA$4, IF(ISNUMBER(実績!$AH${base_row}), 実績!$AH${base_row}, 0), 0))',
    'BI': '=IF($BA{r}="","", BH{r})',
    'BJ': '=IF($BA{r}="","", MIN(BI{r}, SUMIFS($BG$4:$BG$34, $BA$4:$BA$34, $BA{r}, $B$4:$B$34, $BA{r}+5)))',
    'BK': '=IF($BA{r}="","", MAX(0, BI{r} - BJ{r}))',
    'BL': '=IF(OR($BA{r}="", WEEKDAY($B{r},2)>5), 0, SUMIFS($BG$4:$BG$34, $BA$4:$BA$34, $BA{r}, $B$4:$B$34, "<"&$B{r}, $B$4:$B$34, "<="&($BA{r}+4)))',
    # 【安全改修】仕事内容の列に依存せず、出勤状況(D)=休出 かつ 現場(F)=千葉 の場合に集計対象とする
    'BM': '=IF(OR(AND($D{r}="出張", $F{r}<>""), AND($D{r}="休出", $F{r}="千葉")), IF(COUNTIFS($D$4:$D{r}, "出張", $F$4:$F{r}, $F{r}) + COUNTIFS($D$4:$D{r}, "休出", $F$4:$F{r}, "千葉")=1, ROW(), ""), "")',
    'R': '=IF($B{r}="","", IF(ROUND(BE{r}, 5)<=0, "", BE{r}))',
    'S': '=IF($B{r}="","", IF(WEEKDAY($B{r},2)=6, IF(ROUND(MIN(BG{r}, BI{r}), 5)<=0, "", MIN(BG{r}, BI{r})), IF(WEEKDAY($B{r},2)<=5, IF(ROUND(MIN(BG{r}, MAX(0, BK{r} - BL{r})), 5)<=0, "", MIN(BG{r}, MAX(0, BK{r} - BL{r}))), "")))',
    'T': '=IF($B{r}="","", IF(ROUND(BF{r} - N(S{r}), 5)<=0, "", BF{r} - N(S{r})))',
    'U': '=IF(OR($BB{r}="", $BC{r}=""), "", IF(ROUND(MAX(0, MIN($BC{r}, 1+5/24) - MAX($BB{r}, 22/24)) + MAX(0, MIN($BC{r}, 5/24) - $BB{r}), 5)<=0, "", MAX(0, MIN($BC{r}, 1+5/24) - MAX($BB{r}, 22/24)) + MAX(0, MIN($BC{r}, 5/24) - $BB{r})))',
    'V': '=IF($B{r}="","", IF(OR($D{r}="休出", WEEKDAY($B{r},2)=7), IF(OR(BB{r}="", BC{r}=""), "", IF(ROUND(MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))), 5)<=0, "", MAX(0, BC{r} - BB{r} - MAX(0, MIN(BC{r}, TIME(10,30,0)) - MAX(BB{r}, TIME(10,0,0))) - MAX(0, MIN(BC{r}, TIME(13,0,0)) - MAX(BB{r}, TIME(12,0,0))) - MAX(0, MIN(BC{r}, TIME(14,30,0)) - MAX(BB{r}, TIME(14,0,0)))))), ""))',
    'W': '=IF($D{r}="有給", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")',
    'X': '=IF($B{r}="","", IF(OR($D{r}="有給",$D{r}="欠勤",$D{r}="特別",$D{r}="休出",WEEKDAY($B{r},2)=7,AND(BB{r}="",BC{r}="")), "", IF(WEEKDAY($B{r},2)<=5, IF(ROUND(7/24-BE{r}, 5)>0, 7/24-BE{r}, ""), IF(ROUND(5/24-BE{r}, 5)>0, 5/24-BE{r}, ""))))',
    'Y': '=IF($D{r}="欠勤", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")',
    'Z': '=IF($D{r}="特別", IF(WEEKDAY($B{r},2)<=5, 7/24, 5/24), "")'
}

for sheet_name in wb.sheetnames:
    if sheet_name in exclude_sheets: continue
    ws = wb[sheet_name]
    clean_sn = sheet_name.replace(" ", "").replace("　", "")
    base_row = emp_rows.get(clean_sn)
    if not base_row: continue

    # 【自動探知】仕事内容の列をヘッダーから探す
    job_content_col = None
    for c_idx in range(1, 20):
        val = ws.cell(row=3, column=c_idx).value
        if val and ("内容" in str(val) or "仕事" in str(val) or "作業" in str(val) or "備考" in str(val)):
            job_content_col = openpyxl.utils.get_column_letter(c_idx)
            break

    for col in ['AG', 'AH', 'AI', 'AJ', 'AK', 'AL']:
        for row in range(3, 40):
            ws[f'{col}{row}'].value = None
            ws[f'{col}{row}'].fill = PatternFill()
            ws[f'{col}{row}'].border = Border()

    ws['AH1'] = "前月最終週の不足:"
    ws['AH1'].alignment = Alignment(horizontal="right")
    ws['AJ2'] = f"=実績!$AH${base_row}"
    ws['AJ2'].number_format = '[h]:mm'
    ws['AJ2'].fill = input_fill
    ws['AJ2'].border = thin_border

    ws['AI1'] = "当月最終週の不足:"
    ws['AI1'].alignment = Alignment(horizontal="right")
    ws['AJ1'] = '=IF($B$4="","", IF(WEEKDAY(MAX($B$4:$B$34),2)=7, 0, SUMIFS($W$4:$W$34, $BA$4:$BA$34, MAX($BA$4:$BA$34)) + SUMIFS($X$4:$X$34, $BA$4:$BA$34, MAX($BA$4:$BA$34)) + SUMIFS($Y$4:$Y$34, $BA$4:$BA$34, MAX($BA$4:$BA$34)) + SUMIFS($Z$4:$Z$34, $BA$4:$BA$34, MAX($BA$4:$BA$34))))'
    ws['AJ1'].number_format = '[h]:mm'
    ws['AJ1'].fill = auto_fill
    ws['AJ1'].border = thin_border

    ws.column_dimensions['AK'].width = 8
    ws.column_dimensions['AL'].width = 8
    ws['AK3'] = "例外出社"
    ws['AK3'].font = Font(bold=True)
    ws['AK3'].alignment = Alignment(horizontal="center")
    ws['AK3'].fill = input_fill
    ws['AK3'].border = thin_border
    
    ws['AL3'] = "例外退社"
    ws['AL3'].font = Font(bold=True)
    ws['AL3'].alignment = Alignment(horizontal="center")
    ws['AL3'].fill = input_fill
    ws['AL3'].border = thin_border
    
    for row in range(4, 35):
        d_num = row - 3
        c_let = openpyxl.utils.get_column_letter(d_num + 2)
        
        ws[f'D{row}'] = f'=IF(実績!{c_let}{base_row}="","",IF(実績!{c_let}{base_row}="現場(大牟田)","現場",IF(実績!{c_let}{base_row}="現場(工場)","現場",IF(実績!{c_let}{base_row}="出張(千葉)","出張",IF(実績!{c_let}{base_row}="休出(出張)","休出",実績!{c_let}{base_row})))))'
        
        ws[f'F{row}'] = f'=IF(実績!{c_let}{base_row}="","",IF(実績!{c_let}{base_row}="現場(大牟田)","大牟田",IF(実績!{c_let}{base_row}="現場(工場)","工場",IF(実績!{c_let}{base_row}="出張(千葉)","千葉",IF(実績!{c_let}{base_row}="休出(出張)","千葉","")))))'
        
        # もし自動探知で列が見つかっており、かつ結合セルでない場合に限り安全に書き込む
        if job_content_col and type(ws[f'{job_content_col}{row}']).__name__ != 'MergedCell':
            ws[f'{job_content_col}{row}'] = f'=IF(実績!{c_let}{base_row}="休出(出張)","出張","")'
        
        ws[f'AJ{row}'] = f'=実績!{c_let}{base_row+2}'
        ws[f'AJ{row}'].font = Font(color="FFFFFF")
        
        for col, f_template in formulas.items():
            ws[f'{col}{row}'].value = f_template.format(r=row, base_row=base_row)
            
        ws[f'AK{row}'].fill = input_fill
        ws[f'AK{row}'].border = thin_border
        ws[f'AK{row}'].number_format = '[h]:mm'
        
        ws[f'AL{row}'].fill = input_fill
        ws[f'AL{row}'].border = thin_border
        ws[f'AL{row}'].number_format = '[h]:mm'

        if type(ws[f'J{row}']).__name__ != 'MergedCell':
            ws[f'J{row}'].value = f'=IF($AK{row}<>"",$AK{row},IF(OR($D{row}="現場",$D{row}="出張",$D{row}="休出"),8/24,""))'
            ws[f'J{row}'].number_format = '[h]:mm'
        if type(ws[f'P{row}']).__name__ != 'MergedCell':
            ws[f'P{row}'].value = f'=IF($AL{row}<>"",$AL{row},IF(OR($D{row}="現場",$D{row}="出張",$D{row}="休出"),IF(WEEKDAY($B{row},2)=6,15/24,17/24)+IF(ISNUMBER(実績!{c_let}{base_row+1}),実績!{c_let}{base_row+1}/24,0),IF(IF($J{row}<>"",$J{row},$L{row})<>"",IF($J{row}<>"",$J{row},$L{row})+IF(WEEKDAY($B{row},2)=6,7/24,9/24)+IF(ISNUMBER(実績!{c_let}{base_row+1}),実績!{c_let}{base_row+1}/24,0),"")))'
            ws[f'P{row}'].number_format = '[h]:mm'

    for c in ['B','C','D','E','F','G']:
        ws[f'{c}38'].border = Border(bottom=medium_side, left=thin_side, right=thin_side, top=thin_side)

    ws.merge_cells('I37:I38')
    ws['I37'] = "出張"
    ws['I37'].alignment = Alignment(horizontal="center", vertical="center")
    ws['I37'].font = Font(bold=True, size=14)
    for c_idx in range(10, 13):
        c_let = openpyxl.utils.get_column_letter(c_idx)
        k = c_idx - 9
        ws[f'{c_let}37'] = f'=IFERROR(INDEX($F:$F, SMALL($BM$4:$BM$34, {k})), "")'
        ws[f'{c_let}37'].font = Font(size=14)
        ws[f'{c_let}37'].alignment = Alignment(shrink_to_fit=True, wrap_text=True, horizontal="center", vertical="center")
        
        # 【安全改修】仕事内容の列に頼らず、D列とF列だけで正確に出張をカウントする
        ws[f'{c_let}38'] = f'=IF({c_let}37="", "", COUNTIFS($D$4:$D$34, "出張", $F$4:$F$34, {c_let}37) + COUNTIFS($D$4:$D$34, "休出", $F$4:$F$34, "千葉"))'
        ws[f'{c_let}38'].font = Font(size=18)
        ws[f'{c_let}38'].alignment = Alignment(horizontal="center", vertical="center")
        
    for r in [37, 38]:
        for c_idx in range(9, 13):
            cell = ws.cell(row=r, column=c_idx)
            t_l = medium_side if r == 37 else thin_side
            b_l = medium_side if r == 38 else thin_side
            l_l = medium_side if c_idx == 9 else thin_side
            r_l = medium_side if c_idx == 12 else thin_side
            if c_idx == 9:
                if r == 37: b_l = Side(border_style=None)
                if r == 38: t_l = Side(border_style=None)
            cell.border = Border(top=t_l, bottom=b_l, left=l_l, right=r_l)

    for r in range(39, 41):
        for c_idx in range(1, 27):
            cell = ws.cell(row=r, column=c_idx)
            if cell.value is not None:
                cur_align = cell.alignment
                if cur_align:
                    cell.alignment = Alignment(horizontal=cur_align.horizontal, vertical=cur_align.vertical, wrap_text=cur_align.wrap_text, shrink_to_fit=True)
                else:
                    cell.alignment = Alignment(shrink_to_fit=True)

wb.save(output_path)
print(f"完了しました！ '{output_path}' を確認してください。")