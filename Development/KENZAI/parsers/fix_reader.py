"""
fix_reader.py
差異レポートExcelを読み込み、ユーザーが手動で入力した修正値を取得する。
「修正モード」が「手動」または「システム」の行を抽出し、corrections.py と同じ形式の辞書を生成する。
"""

import openpyxl
from collections import defaultdict
from typing import Dict, Any

def _parse_time_val(val: Any) -> float:
    """HH:MM 形式の文字列または datetime.time オブジェクトを hours (float) に変換する"""
    if val is None:
        return None
    from datetime import time
    if isinstance(val, time):
        return val.hour + val.minute / 60.0
    if isinstance(val, str):
        if ':' in val:
            parts = val.split(':')
            try:
                return int(parts[0]) + int(parts[1]) / 60.0
            except:
                return None
        else:
            try:
                return float(val)
            except:
                return None
    if isinstance(val, (int, float)):
        return float(val)
    return None

def read_fixes(filepath: str) -> Dict[str, Dict[int, Dict[str, Any]]]:
    """
    差異レポートを読み込み、以下の形式の辞書を返す:
    {
        '社員名': {
            日(int): { 'フィールド名': 値, ... }
        }
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    corrections = defaultdict(lambda: defaultdict(dict))

    # ヘッダーを探す（通常5行目だが、柔軟に）
    header_row = 5
    # 列インデックスの特定
    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            col_map[cell.value] = cell.column

    # 必須列の存在確認
    required_cols = ['氏名', '日付', '修正モード', '修正フィールド', '修正値']
    for req in required_cols:
        if req not in col_map:
            raise ValueError(f"差異レポートに必須列 '{req}' が見つかりません。")

    # フィールド名マッピング
    FIELD_MAP = {
        '出社時間': 't_start',
        '退社時間': 't_end',
        '現場開始時間': 't_site_s',
        '現場終了時間': 't_site_e',
        '勤務時間': 'excel_work',
        '欠勤時間': 'excel_absence',
        '休憩(分)': 'break_mins',
        '所内残業': 'excel_ot_in',
        '法外残業': 'excel_ot_out',
        '休出': 'excel_holiday_w',
        '遅刻早退時間': 'excel_absence', # 現状、欠勤時間と同じ扱い
        '有給（時間、半日、1日）': 'time_off',
    }

    current_emp = None

    for row in range(header_row + 1, ws.max_row + 1):
        emp_name = ws.cell(row=row, column=col_map['氏名']).value
        if emp_name:
            current_emp = str(emp_name)

        if not current_emp:
            continue

        mode = ws.cell(row=row, column=col_map['修正モード']).value
        if not mode or mode == '':
            continue

        date_str = ws.cell(row=row, column=col_map['日付']).value
        if not date_str:
            continue

        # "1日" から 1 を抽出
        try:
            day_num = int(str(date_str).replace('日', '').strip())
        except ValueError:
            continue

        if mode == 'システム':
            # システム計算値を採用する場合、特に上書きは不要（計算結果がそのまま出力されるため）
            # ただし「確認済み」フラグとして持たせることもできるが、今回はスキップ
            continue

        if mode == '手動':
            field_raw = ws.cell(row=row, column=col_map['修正フィールド']).value
            val_raw = ws.cell(row=row, column=col_map['修正値']).value

            if not field_raw or field_raw not in FIELD_MAP:
                continue

            dict_key = FIELD_MAP[field_raw]
            
            # 値のパース
            if field_raw in ['出社時間', '退社時間', '現場開始時間', '現場終了時間']:
                parsed_val = _parse_time_val(val_raw)
            elif field_raw == '休憩(分)':
                try:
                    parsed_val = float(val_raw)
                except:
                    parsed_val = 0.0
            else:
                try:
                    parsed_val = float(val_raw)
                except:
                    parsed_val = 0.0
            
            if parsed_val is not None:
                corrections[current_emp][day_num][dict_key] = parsed_val

    # defaultdictを通常のdictに変換
    return {k: dict(v) for k, v in corrections.items()}
