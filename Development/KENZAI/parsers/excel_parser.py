"""
excel_parser.py
Excel出勤簿パーサー（平野工業用）。
現行 payroll_engine.py の ExcelParser を移植・リファクタリング。
CompanyConfig を参照しつつ、後方互換のdict形式でデータを返す。
"""

import openpyxl
from datetime import datetime, time
from typing import List, Dict, Any

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from parsers.base_parser import InputParser
from base_config import CompanyConfig


# ── Excel列マッピング（平野工業出勤簿の固定レイアウト） ──
COL = {
    'date':      2,   # B: 日付
    'weekday':   3,   # C: 曜日
    'place':     4,   # D: 工事場所
    'site_name': 6,   # F: 現場名
    'client':    8,   # H: 客先名
    'depart':    10,  # J: 出社時間
    'site_s':    12,  # L: 現場開始時間
    'site_e':    14,  # N: 現場終了時間
    'arrive':    16,  # P: 退社時間
    'work':      18,  # R: 勤務時間
    'ot_out':    19,  # S: 法定外残業
    'ot_in':     20,  # T: 所内（法定内残業）
    'holiday_w': 21,  # U: 休出
    'absence':   22,  # V: 欠勤
    'time_off':  23,  # W: 時間休
    'task':      24,  # X: 作業内容
}

SUMMARY_ROW = 35
LABEL_ROW   = 37
PAID_H_ROW  = 38


# ── ユーティリティ ──

def _to_hours(val):
    """time/datetime/str/None → float(時間数)。変換不能はNone。"""
    if val is None:
        return None
    if isinstance(val, time):
        return val.hour + val.minute / 60.0
    if isinstance(val, datetime):
        return val.hour + val.minute / 60.0
    if isinstance(val, str):
        parts = val.split(':')
        try:
            return int(parts[0]) + int(parts[1]) / 60.0
        except Exception:
            return None
    return None


def _to_num(val):
    """任意の値をfloatへ。変換不能は0.0。"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except Exception:
            return 0.0
    return 0.0


class ExcelParser(InputParser):
    """出勤簿Excelを読み取り、従業員ごとの月次データを返す。"""

    def __init__(self, config: CompanyConfig):
        super().__init__(config)
        self._wb = None
        self._filepath = None

    def parse(self, source_path: str, employee_master: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Excelファイルを全シート解析し、社員ごとのデータリストを返す。

        Args:
            source_path: Excelファイルパス
            employee_master: 社員マスター辞書 {シート名: {code, name}}

        Returns:
            list of dict
        """
        self._filepath = source_path
        self._wb = openpyxl.load_workbook(source_path, data_only=True)

        results = []
        for sheet_name in self._wb.sheetnames:
            if sheet_name not in employee_master:
                continue
            data = self._parse_sheet(sheet_name)
            data['is_special'] = self.config.is_special_employee(sheet_name)
            results.append(data)

        return results

    def get_sheet_names(self):
        """シート名のリストを返す"""
        if self._wb is None:
            return []
        return self._wb.sheetnames

    def parse_sheet(self, sheet_name):
        """
        後方互換: 1シート分を解析して Dict を返す。
        """
        return self._parse_sheet(sheet_name)

    def _parse_sheet(self, sheet_name):
        """1シート分を解析して Dict を返す。"""
        ws = self._wb[sheet_name]
        is_special = self.config.is_special_employee(sheet_name)

        year, month = self._detect_year_month(ws)

        days = []
        for rn in range(4, SUMMARY_ROW):
            rec = self._parse_day_row(ws, rn, sheet_name, is_special)
            if rec is not None:
                days.append(rec)

        summary = self._parse_summary(ws)

        return {
            'sheet_name': sheet_name,
            'year': year,
            'month': month,
            'hirano': is_special,       # 後方互換
            'is_special': is_special,
            'days': days,
            'summary_excel': summary,
        }

    def _detect_year_month(self, ws):
        """シートから年月を検出する。"""
        for rn in range(4, 35):
            d = ws.cell(row=rn, column=COL['date']).value
            if isinstance(d, datetime):
                return d.year, d.month
        return None, None

    def _parse_day_row(self, ws, row_num, sheet_name, is_special):
        """1行分の日次データを解析して dict を返す。空行はNone。"""
        get = lambda col_key: ws.cell(row=row_num, column=COL[col_key]).value

        date_v = get('date')
        weekday = get('weekday')
        place = get('place')

        if not isinstance(date_v, datetime):
            return None

        day_num = date_v.day

        # 時刻
        t_depart = _to_hours(get('depart'))
        t_site_s = _to_hours(get('site_s'))
        t_site_e = _to_hours(get('site_e'))
        t_arrive = _to_hours(get('arrive'))

        # 実績開始・終了（優先順: J>L, P>N）
        t_start = t_depart if t_depart is not None else t_site_s
        t_end = t_arrive if t_arrive is not None else t_site_e

        # Excel記載値（集計検証用）
        excel_work = _to_num(get('work'))
        excel_ot_out = _to_num(get('ot_out'))
        excel_ot_in = _to_num(get('ot_in'))
        excel_holiday_w = _to_num(get('holiday_w'))
        excel_absence = _to_num(get('absence'))
        time_off_raw = get('time_off')

        # 時間休の処理（半日 = CompanyConfigから取得）
        if time_off_raw == '半日':
            time_off = self.config.get_half_day_hours(sheet_name)
        else:
            time_off = _to_num(time_off_raw)

        # 場所の判定
        is_absent = (str(place) == '欠') if place else False
        is_paid = (str(place) == '有給') if place else False
        is_training = (isinstance(place, str) and '研修' in place)
        is_holiday = weekday in ('土', '日')

        return {
            'row': row_num,
            'day': day_num,
            'weekday': weekday,
            'place': str(place) if place else '',
            't_start': t_start,
            't_end': t_end,
            't_depart': t_depart,
            't_site_s': t_site_s,
            't_site_e': t_site_e,
            't_arrive': t_arrive,
            'time_off': time_off,
            'time_off_raw': str(time_off_raw) if time_off_raw else '',
            'is_absent': is_absent,
            'is_paid': is_paid,
            'is_training': is_training,
            'is_saturday': weekday == '土',
            'is_sunday': weekday == '日',
            # Excel記載値（後で計算値と突合するため保持）
            'excel_work': excel_work,
            'excel_ot_out': excel_ot_out,
            'excel_ot_in': excel_ot_in,
            'excel_holiday_w': excel_holiday_w,
            'excel_absence': excel_absence,
        }

    def _parse_summary(self, ws):
        """集計行(35/37/38)を辞書で返す（検証用）。"""
        r35 = {c.column: c.value for c in ws[SUMMARY_ROW] if c.value is not None}
        r37 = {c.column: c.value for c in ws[LABEL_ROW] if c.value is not None}
        try:
            r38 = {c.column: c.value for c in ws[PAID_H_ROW] if c.value is not None}
        except Exception:
            r38 = {}
        return {
            'total_work': _to_num(r35.get(COL['work'], 0)),
            'total_ot_out': _to_num(r35.get(COL['ot_out'], 0)),
            'total_ot_in': _to_num(r35.get(COL['ot_in'], 0)),
            'total_absence': _to_num(r35.get(COL['absence'], 0)),
            'total_time_off': _to_num(r35.get(COL['time_off'], 0)),
            'attend_days': _to_num(r37.get(3, 0)),
            'paid_leave_days': _to_num(r37.get(25, 0)),
            'paid_leave_hours': _to_num(r38.get(25, 0)),
        }

    def load_workbook(self, filepath):
        """Excelファイルを読み込む（後方互換用）"""
        self._filepath = filepath
        self._wb = openpyxl.load_workbook(filepath, data_only=True)
