"""
excel_clone_exporter.py
入力Excelと同じフォーマットで計算結果を出力するエクスポーター。

入力のExcelファイルをコピーし、計算結果で勤務時間関連の列（R〜W列 + 集計行）を
上書きする方式。これにより、ヘッダー・レイアウト・マージセル・書式は完全保持される。

対象列（上書き箇所）:
  R列（18）: 勤務時間
  S列（19）: 残業（法定外残業）
  T列（20）: 所内（延長時間 = 法定内残業）
  U列（21）: 休出
  V列（22）: 欠勤（遅刻早退時間）
  W列（23）: 時間休 → 変更なし（入力値を維持）
  35行目: 月次集計行
  37行目: サマリーラベル行（出勤日数等を更新）
"""

import os
import shutil
import openpyxl

# 入力Excelの列マッピング（excel_parser.py と同一）
COL = {
    'work':      18,  # R: 勤務時間
    'ot_out':    19,  # S: 残業（法定外残業）
    'ot_in':     20,  # T: 所内（法定内残業）
    'holiday_w': 21,  # U: 休出
    'absence':   22,  # V: 欠勤（遅刻早退）
    'time_off':  23,  # W: 時間休
}

SUMMARY_ROW = 35
LABEL_ROW   = 37
PAID_H_ROW  = 38


class ExcelCloneExporter:
    """
    入力Excelをコピーし、計算結果で上書き出力するエクスポーター。
    書式・レイアウト・マージセルは完全保持される。
    """

    def export(self, input_path, output_path, all_employee_results):
        """
        入力Excelをコピーし、計算結果で勤務時間列を上書きする。

        Args:
            input_path: 入力Excelファイルパス（元ファイル）
            output_path: 出力先パス（コピーして上書き）
            all_employee_results: list of dict
                [{
                    'sheet_name': str（シート名 = 社員名、入力Excelのシート名と一致）,
                    'calc_results': [(day_rec, calc), ...],
                    'monthly': dict（月次集計結果）,
                }]
        """
        # 入力ファイルをコピー（書式・マージセル・印刷設定を完全保持）
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        shutil.copy2(input_path, output_path)

        # コピーしたファイルを開いて計算結果で上書き
        wb = openpyxl.load_workbook(output_path)

        updated_count = 0
        for emp_data in all_employee_results:
            sheet_name = emp_data['sheet_name']
            if sheet_name not in wb.sheetnames:
                print(f"  [スキップ] シート '{sheet_name}' が入力Excelに存在しません")
                continue

            ws = wb[sheet_name]
            calc_results = emp_data['calc_results']
            monthly = emp_data['monthly']

            self._update_daily_rows(ws, calc_results)
            self._update_summary_row(ws, monthly)
            self._update_label_row(ws, monthly)

            updated_count += 1

        wb.save(output_path)
        print(f"[出勤簿出力] 完了: {output_path} ({updated_count}名分)")
        return output_path

    def _update_daily_rows(self, ws, calc_results):
        """日次データ行（4〜34行）の勤務時間列を計算結果で上書きする。"""
        for day_rec, calc in calc_results:
            row_num = day_rec['row'] if isinstance(day_rec, dict) else day_rec.row

            # dict/dataclass 両対応
            calc_work = calc['work'] if isinstance(calc, dict) else calc.work
            calc_ot_out = calc['ot_out'] if isinstance(calc, dict) else calc.ot_out
            calc_ot_in = calc['ot_in'] if isinstance(calc, dict) else calc.ot_in
            calc_absence = calc['absence'] if isinstance(calc, dict) else calc.absence

            is_paid = day_rec['is_paid'] if isinstance(day_rec, dict) else day_rec.is_paid
            is_absent = day_rec['is_absent'] if isinstance(day_rec, dict) else day_rec.is_absent
            is_sunday = day_rec['is_sunday'] if isinstance(day_rec, dict) else day_rec.is_sunday

            # R列: 勤務時間（有給・欠勤の日は0）
            ws.cell(row=row_num, column=COL['work']).value = (
                calc_work if calc_work > 0 else None
            )

            # S列: 法定外残業（残業）
            ws.cell(row=row_num, column=COL['ot_out']).value = (
                calc_ot_out if calc_ot_out > 0 else None
            )

            # T列: 所内（法定内残業 = 延長時間）
            ws.cell(row=row_num, column=COL['ot_in']).value = (
                calc_ot_in if calc_ot_in > 0 else None
            )

            # U列: 休出（日曜出勤の勤務時間）
            holiday_work = 0
            if is_sunday and calc_work > 0:
                holiday_work = calc_work
            ws.cell(row=row_num, column=COL['holiday_w']).value = (
                holiday_work if holiday_work > 0 else None
            )

            # V列: 遅刻早退（欠勤時間）
            ws.cell(row=row_num, column=COL['absence']).value = (
                calc_absence if calc_absence > 0 else None
            )

            # W列（時間休）は入力値を維持 → 変更しない

    def _update_summary_row(self, ws, monthly):
        """集計行（35行目）を計算結果で上書きする。"""
        ws.cell(row=SUMMARY_ROW, column=COL['work']).value = monthly['total_work']
        ws.cell(row=SUMMARY_ROW, column=COL['ot_out']).value = monthly['total_ot_out']
        ws.cell(row=SUMMARY_ROW, column=COL['ot_in']).value = monthly['total_ot_in']
        ws.cell(row=SUMMARY_ROW, column=COL['holiday_w']).value = monthly.get('total_holiday_w', 0)
        ws.cell(row=SUMMARY_ROW, column=COL['absence']).value = monthly['total_absence']

    def _update_label_row(self, ws, monthly):
        """サマリーラベル行（37行目）の出勤日数等を更新する。"""
        # (3)出勤日数
        ws.cell(row=LABEL_ROW, column=3).value = monthly['attend_days']

        # (5)勤務合計
        ws.cell(row=LABEL_ROW, column=5).value = monthly['total_work']

        # (9)休出合計
        ws.cell(row=LABEL_ROW, column=9).value = monthly.get('total_holiday_w', 0)

        # (13)残業合計
        ws.cell(row=LABEL_ROW, column=13).value = monthly['total_ot_out']

        # (17)所内合計
        ws.cell(row=LABEL_ROW, column=17).value = monthly['total_ot_in']

        # (21)欠勤合計
        ws.cell(row=LABEL_ROW, column=21).value = monthly['total_absence']

        # (25)有給日数
        ws.cell(row=LABEL_ROW, column=25).value = monthly['paid_days']

        # (25, 38行目)有給時間
        ws.cell(row=PAID_H_ROW, column=25).value = monthly['paid_hours']
