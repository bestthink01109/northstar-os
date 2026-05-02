"""
diff_report.py
入力Excelと計算エンジンの結果を比較し、差異レポートファイルを出力する。

出力フォーマット: Excel（.xlsx）
  - シート1枚に全社員の差異を一覧表示
  - 氏名、日付、項目、Excel記載値、計算値、差分を明記
  - 差異がない社員は「差異なし」と表示
  - 要確認アラートは黄色背景で表示（出勤簿を手修正 → 再実行で解消）
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


class DiffReportExporter:
    """入力Excel vs 計算結果 の差異レポートを生成する。"""

    # スタイル定義
    HEADER_FONT = Font(name='游ゴシック', bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    OK_FILL     = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    NG_FILL     = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    ALERT_FILL  = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # 要確認アラート：黄色
    DATA_FONT   = Font(name='游ゴシック', size=10)
    BOLD_FONT   = Font(name='游ゴシック', bold=True, size=10)
    NAME_FONT   = Font(name='游ゴシック', bold=True, size=11)
    ALERT_FONT  = Font(name='游ゴシック', bold=True, size=10, color='7D6608')  # アラート文字：濃い黄色
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT   = Alignment(horizontal='left', vertical='center')

    HEADERS = ['氏名', '社員コード', '日付', '曜日', '項目',
               'Excel記載値', '計算値', '差分', '備考',
               '修正モード', '修正フィールド', '修正値', '理由']

    def export(self, diff_data, output_path, company_name, year, month):
        """
        差異レポートを出力する。

        Args:
            diff_data: list of dict
                [{
                    'employee_name': str,
                    'employee_code': int,
                    'issues': list of dict（日次差異の詳細）,
                    'summary_issues': list of dict（月次集計差異の詳細）,
                }]
            output_path: 出力先パス
            company_name: 会社名
            year, month: 対象年月
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '差異レポート'

        # ── タイトル ──
        ws.merge_cells('A1:M1')
        title = ws['A1']
        title.value = f"{company_name}　差異レポート　{year}年{month}月"
        title.font = Font(name='游ゴシック', bold=True, size=14)
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # ── サマリー行（2行目） ──
        total     = len(diff_data)
        has_diff  = sum(1 for d in diff_data if d['issues'] or d['summary_issues'])
        no_diff   = total - has_diff
        has_alert = sum(
            1 for d in diff_data
            if any(i.get('is_alert') for i in d['issues'])
        )
        ws.merge_cells('A2:M2')
        summary = ws['A2']
        alert_text = f"　要確認アラート（黄色）: {has_alert}名" if has_alert else ''
        summary.value = (
            f"対象: {total}名　差異あり: {has_diff}名　差異なし: {no_diff}名{alert_text}"
        )
        summary.font = Font(name='游ゴシック', size=11)
        summary.alignment = self.CENTER

        # ── 凡例行（3行目） ──
        ws.merge_cells('A3:M3')
        legend = ws['A3']
        legend.value = (
            '【凡例】赤背景 = 計算差異（要調査）　黄背景 = 要確認アラート'
            '（出勤簿を手修正 → 再実行で解消）'
        )
        legend.font = Font(name='游ゴシック', size=9, color='555555')
        legend.alignment = self.LEFT

        # ── ヘッダー行（5行目） ──
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER
            cell.border = self.THIN_BORDER

        # 列幅設定（備考列を幅広に）
        col_widths = [16, 10, 10, 6, 18, 12, 12, 10, 42, 12, 24, 10, 30]
        for i, w in enumerate(col_widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── データ行（6行目〜） ──
        row = 6
        for emp in diff_data:
            emp_name       = emp['employee_name']
            emp_code       = emp['employee_code']
            issues         = emp.get('issues', [])
            summary_issues = emp.get('summary_issues', [])

            if not issues and not summary_issues:
                # 差異なしの社員
                cell_name = ws.cell(row=row, column=1, value=emp_name)
                cell_name.font      = self.BOLD_FONT
                cell_name.alignment = self.LEFT
                cell_name.border    = self.THIN_BORDER

                cell_code = ws.cell(row=row, column=2, value=emp_code)
                cell_code.font      = self.DATA_FONT
                cell_code.alignment = self.CENTER
                cell_code.border    = self.THIN_BORDER

                ws.merge_cells(start_row=row, start_column=3,
                               end_row=row, end_column=13)
                ok_cell = ws.cell(row=row, column=3, value='差異なし')
                ok_cell.font      = self.BOLD_FONT
                ok_cell.alignment = self.CENTER
                ok_cell.border    = self.THIN_BORDER
                for cn in range(3, 14):
                    ws.cell(row=row, column=cn).fill   = self.OK_FILL
                    ws.cell(row=row, column=cn).border = self.THIN_BORDER
                row += 1

            else:
                # 日次差異
                first_row = True
                for issue in issues:
                    is_alert = issue.get('is_alert', False)

                    c1 = ws.cell(row=row, column=1,
                                 value=emp_name if first_row else '')
                    c1.font      = self.NAME_FONT if first_row else self.DATA_FONT
                    c1.alignment = self.LEFT
                    c1.border    = self.THIN_BORDER

                    c2 = ws.cell(row=row, column=2,
                                 value=emp_code if first_row else '')
                    c2.font      = self.DATA_FONT
                    c2.alignment = self.CENTER
                    c2.border    = self.THIN_BORDER

                    vals = [
                        issue.get('date_str', ''),
                        issue.get('weekday', ''),
                        issue.get('item', ''),
                        issue.get('excel_val', ''),
                        issue.get('calc_val', ''),
                        issue.get('diff', ''),
                        issue.get('note', ''),
                        '',  # 修正モード
                        '',  # 修正フィールド
                        '',  # 修正値
                        '',  # 理由
                    ]
                    fill = self.ALERT_FILL if is_alert else self.NG_FILL
                    font = self.ALERT_FONT if is_alert else self.DATA_FONT

                    for ci, v in enumerate(vals, 3):
                        cell = ws.cell(row=row, column=ci, value=v)
                        cell.font      = font
                        cell.alignment = self.LEFT if ci == 9 else self.CENTER
                        cell.border    = self.THIN_BORDER
                        cell.fill      = fill

                    first_row = False
                    row += 1

                # 月次集計差異
                for s_issue in summary_issues:
                    c1 = ws.cell(row=row, column=1,
                                 value=emp_name if first_row else '')
                    c1.font      = self.BOLD_FONT if first_row else self.DATA_FONT
                    c1.alignment = self.LEFT
                    c1.border    = self.THIN_BORDER

                    c2 = ws.cell(row=row, column=2,
                                 value=emp_code if first_row else '')
                    c2.font      = self.DATA_FONT
                    c2.alignment = self.CENTER
                    c2.border    = self.THIN_BORDER

                    vals = [
                        '月次集計',
                        '',
                        s_issue.get('item', ''),
                        s_issue.get('excel_val', ''),
                        s_issue.get('calc_val', ''),
                        s_issue.get('diff', ''),
                        s_issue.get('note', ''),
                        '',  # 修正モード
                        '',  # 修正フィールド
                        '',  # 修正値
                        '',  # 理由
                    ]
                    for ci, v in enumerate(vals, 3):
                        cell = ws.cell(row=row, column=ci, value=v)
                        cell.font      = self.BOLD_FONT
                        cell.alignment = self.CENTER
                        cell.border    = self.THIN_BORDER
                        cell.fill      = self.NG_FILL

                    first_row = False
                    row += 1

            # 社員間に空行
            row += 1

        # データ入力規則（ドロップダウン）の設定
        # 修正モード: J列 (col 10)
        mode_dv = DataValidation(type="list", formula1='"システム,手動"', allow_blank=True)
        mode_dv.error = '無効な値です'
        mode_dv.errorTitle = '入力エラー'
        ws.add_data_validation(mode_dv)
        mode_dv.add(f'J6:J{row}')

        # 修正フィールド: K列 (col 11)
        fields = "出社時間,退社時間,現場開始時間,現場終了時間,勤務時間,欠勤時間,休憩(分),所内残業,休出,法外残業,遅刻早退時間,有給（時間、半日、1日）"
        field_dv = DataValidation(type="list", formula1=f'"{fields}"', allow_blank=True)
        field_dv.error = 'リストから選択してください'
        field_dv.errorTitle = '入力エラー'
        ws.add_data_validation(field_dv)
        field_dv.add(f'K6:K{row}')

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"[差異レポート] 完了: {output_path}")
        print(f"  差異あり: {has_diff}名 / 差異なし: {no_diff}名")
        if has_alert:
            print(
                f"  要確認アラート（黄色）: {has_alert}名"
                f" ← 出勤簿を手修正して再実行してください"
            )
        return output_path


def build_diff_data(sheet_name, emp_code, emp_full_name,
                    days, calc_results, monthly, excel_summary):
    """
    1社員分の差異データを構造化して返す。
    DiffReportExporter.export() に渡す形式の dict を生成する。

    is_alert=True の差異は「要確認アラート（黄色）」として表示される。
    対応方法: 入力の出勤簿 Excel を手修正 → main.py を再実行。
    """
    issues         = []
    summary_issues = []

    for day_rec, calc in calc_results:
        d  = day_rec['day']     if isinstance(day_rec, dict) else day_rec.day
        wd = day_rec['weekday'] if isinstance(day_rec, dict) else day_rec.weekday

        calc_work    = calc['work']    if isinstance(calc, dict) else calc.work
        calc_ot_in   = calc['ot_in']   if isinstance(calc, dict) else calc.ot_in
        calc_ot_out  = calc['ot_out']  if isinstance(calc, dict) else calc.ot_out
        calc_absence = calc['absence'] if isinstance(calc, dict) else calc.absence

        excel_work    = day_rec['excel_work']    if isinstance(day_rec, dict) else day_rec.excel_work
        excel_ot_in   = day_rec['excel_ot_in']   if isinstance(day_rec, dict) else day_rec.excel_ot_in
        excel_ot_out  = day_rec['excel_ot_out']  if isinstance(day_rec, dict) else day_rec.excel_ot_out
        excel_absence = day_rec['excel_absence'] if isinstance(day_rec, dict) else day_rec.excel_absence

        # 土曜 + time_off がある日 → アラート判定に使用
        is_sat   = (day_rec.get('is_saturday', False)
                    if isinstance(day_rec, dict)
                    else getattr(day_rec, 'is_saturday', False))
        time_off = (day_rec.get('time_off', 0.0)
                    if isinstance(day_rec, dict)
                    else getattr(day_rec, 'time_off', 0.0))

        date_str = f"{d}日"

        # 勤務時間差異
        if abs(calc_work - excel_work) > 0.05:
            issues.append({
                'date_str': date_str, 'weekday': wd, 'item': '勤務時間',
                'excel_val': f"{excel_work:.1f}h",
                'calc_val':  f"{calc_work:.1f}h",
                'diff':      f"{calc_work - excel_work:+.1f}h",
                'note':      '',
                'is_alert':  False,
            })

        # 所内（法定内残業）差異
        if abs(calc_ot_in - excel_ot_in) > 0.05:
            issues.append({
                'date_str': date_str, 'weekday': wd, 'item': '所内(法定内残業)',
                'excel_val': f"{excel_ot_in:.1f}h",
                'calc_val':  f"{calc_ot_in:.1f}h",
                'diff':      f"{calc_ot_in - excel_ot_in:+.1f}h",
                'note':      '',
                'is_alert':  False,
            })

        # 法定外残業差異
        if abs(calc_ot_out - excel_ot_out) > 0.05:
            issues.append({
                'date_str': date_str, 'weekday': wd, 'item': '法定外残業',
                'excel_val': f"{excel_ot_out:.1f}h",
                'calc_val':  f"{calc_ot_out:.1f}h",
                'diff':      f"{calc_ot_out - excel_ot_out:+.1f}h",
                'note':      '',
                'is_alert':  False,
            })

        # 欠勤差異（土曜+time_off の場合は要確認アラート）
        if abs(calc_absence - excel_absence) > 0.05:
            is_alert = bool(is_sat and time_off > 0.0)
            note = (
                f"要確認：出勤簿{d}日({wd})の時間休・退勤時刻を確認し、"
                f"必要なら出勤簿を修正して再実行してください"
            ) if is_alert else ''
            issues.append({
                'date_str': date_str, 'weekday': wd, 'item': '欠勤',
                'excel_val': f"{excel_absence:.1f}h",
                'calc_val':  f"{calc_absence:.1f}h",
                'diff':      f"{calc_absence - excel_absence:+.1f}h",
                'note':      note,
                'is_alert':  is_alert,
            })

    # 月次集計との突合
    if isinstance(excel_summary, dict):
        checks = [
            ('勤務合計',     'total_work'),
            ('法定外残業合計', 'total_ot_out'),
            ('所内合計',     'total_ot_in'),
            ('遅刻早退合計',  'total_absence'),
        ]
        for label, key in checks:
            calc_v  = monthly.get(key, 0)
            excel_v = excel_summary.get(key, 0)
            if abs(calc_v - excel_v) > 0.05:
                summary_issues.append({
                    'item':      label,
                    'excel_val': f"{excel_v:.1f}h",
                    'calc_val':  f"{calc_v:.1f}h",
                    'diff':      f"{calc_v - excel_v:+.1f}h",
                    'note':      '',
                })

    return {
        'employee_name':  emp_full_name,
        'employee_code':  emp_code,
        'issues':         issues,
        'summary_issues': summary_issues,
    }
