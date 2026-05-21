"""
attendance_sheet.py
出勤簿Excel自動生成モジュール。
全時間データ（出退勤、勤務時間、延長時間、残業時間、遅刻早退、欠勤、有給）を
網羅した出勤簿をExcel形式で自動生成する。
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


class AttendanceSheetExporter:
    """出勤簿Excel自動生成エクスポーター。"""

    # ヘッダー定義（1行目に出力される列名）
    HEADERS = [
        '日', '曜日', '出勤', '退勤', '休憩', '実働',
        '所定', '勤務', '延長', '残業', '遅刻早退',
        '欠勤', '有給種別', '有給時間', '外勤', '備考'
    ]

    # スタイル定義
    HEADER_FONT = Font(name='游ゴシック', bold=True, size=10)
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT_WHITE = Font(name='游ゴシック', bold=True, size=10, color='FFFFFF')
    DATA_FONT = Font(name='游ゴシック', size=10)
    TOTAL_FONT = Font(name='游ゴシック', bold=True, size=10)
    TOTAL_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    SATURDAY_FILL = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    SUNDAY_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def export(self, all_employee_data, output_path, company_name, year, month):
        """
        全社員の出勤簿をExcelファイルとして出力する。

        Args:
            all_employee_data: list of dict
                各要素: {
                    'employee_name': str,
                    'employee_code': int,
                    'calc_results': [(day_rec, calc), ...],
                    'monthly': {...},
                    'is_special': bool,
                }
            output_path: 出力先ファイルパス
            company_name: 会社名
            year, month: 対象年月
        """
        wb = openpyxl.Workbook()
        # デフォルトシートを削除
        wb.remove(wb.active)

        for emp_data in all_employee_data:
            self._create_employee_sheet(wb, emp_data, company_name, year, month)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"[出勤簿] 完了: {output_path} ({len(all_employee_data)}名分)")
        return output_path

    def _create_employee_sheet(self, wb, emp_data, company_name, year, month):
        # 外勤手当情報（純青固有。ない場合は0）
        field_work_allowance_per_day = emp_data.get('field_work_allowance_per_day', 0)
        field_work_total = emp_data.get('field_work_total', 0)
        """1社員分のシートを作成する。"""
        emp_name = emp_data['employee_name']
        emp_code = emp_data['employee_code']
        calc_results = emp_data['calc_results']
        monthly = emp_data['monthly']

        # シート名（Excelの制限31文字以内）
        sheet_title = emp_name[:31]
        ws = wb.create_sheet(title=sheet_title)

        # ── タイトル行 ──
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"{company_name}　出勤簿　{year}年{month}月　{emp_name}（コード:{emp_code}）"
        title_cell.font = Font(name='游ゴシック', bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # ── ヘッダー行（3行目） ──
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

        # ── 列幅設定 ──
        col_widths = [4, 5, 7, 7, 5, 6, 6, 6, 6, 6, 8, 6, 8, 8, 5, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── データ行（4行目〜） ──
        row_num = 4
        for day_rec, calc in calc_results:
            # dict/dataclass 両対応
            day = day_rec['day'] if isinstance(day_rec, dict) else day_rec.day
            weekday = day_rec['weekday'] if isinstance(day_rec, dict) else day_rec.weekday
            t_start = day_rec['t_start'] if isinstance(day_rec, dict) else day_rec.t_start
            t_end = day_rec['t_end'] if isinstance(day_rec, dict) else day_rec.t_end
            time_off_raw = day_rec.get('time_off_raw', '') if isinstance(day_rec, dict) else getattr(day_rec, 'time_off_raw', '')
            time_off = day_rec['time_off'] if isinstance(day_rec, dict) else day_rec.time_off
            is_saturday = day_rec['is_saturday'] if isinstance(day_rec, dict) else day_rec.is_saturday
            is_sunday = day_rec['is_sunday'] if isinstance(day_rec, dict) else day_rec.is_sunday
            is_paid = day_rec['is_paid'] if isinstance(day_rec, dict) else day_rec.is_paid
            is_absent = day_rec['is_absent'] if isinstance(day_rec, dict) else day_rec.is_absent

            calc_work = calc['work'] if isinstance(calc, dict) else calc.work
            calc_ot_in = calc['ot_in'] if isinstance(calc, dict) else calc.ot_in
            calc_ot_out = calc['ot_out'] if isinstance(calc, dict) else calc.ot_out
            calc_absence = calc['absence'] if isinstance(calc, dict) else calc.absence
            calc_actual = calc['actual_work'] if isinstance(calc, dict) else calc.actual_work
            calc_scheduled = calc['scheduled'] if isinstance(calc, dict) else calc.scheduled
            calc_break = calc.get('break_hours', 0) if isinstance(calc, dict) else getattr(calc, 'break_hours', 0)
            # 外勤フラグ（純青固有）
            field_work = day_rec.get('field_work', False) if isinstance(day_rec, dict) else getattr(day_rec, 'field_work', False)

            # 有給種別の判定
            paid_type = ''
            paid_hours = 0.0
            if is_paid:
                paid_type = '全休'
                paid_hours = calc_scheduled
            elif time_off_raw == '半日':
                paid_type = '半日'
                paid_hours = time_off
            elif time_off > 0:
                paid_type = '時間休'
                paid_hours = time_off

            # データ書き込み
            values = [
                day,                                                                # 日
                weekday,                                                            # 曜日
                self._format_time(t_start) if t_start and not is_paid else '',     # 出勤
                self._format_time(t_end) if t_end and not is_paid else '',         # 退勤
                calc_break if calc_break > 0 else '',                              # 休憩
                round(calc_actual, 2) if calc_actual > 0 else '',                  # 実働
                calc_scheduled if calc_scheduled > 0 else '',                      # 所定
                calc_work if calc_work > 0 else '',                                # 勤務
                calc_ot_in if calc_ot_in > 0 else '',                              # 延長
                calc_ot_out if calc_ot_out > 0 else '',                            # 残業
                calc_absence if calc_absence > 0 else '',                          # 遅刻早退
                calc_scheduled if is_absent else '',                               # 欠勤
                paid_type,                                                         # 有給種別
                paid_hours if paid_hours > 0 else '',                              # 有給時間
                '◯' if field_work else '',                                         # 外勤
                '',                                                                # 備考
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = self.DATA_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.THIN_BORDER

                # 土日の行カラーリング
                if is_saturday:
                    cell.fill = self.SATURDAY_FILL
                elif is_sunday:
                    cell.fill = self.SUNDAY_FILL

            row_num += 1

        # ── 集計行 ──
        row_num += 1  # 1行空ける
        # 外勤手当の集計文字列
        field_work_days = monthly.get('total_field_work_days', 0)
        if field_work_days > 0 and field_work_allowance_per_day > 0:
            field_work_str = f"{field_work_days}日"
            field_work_note = f"外勤手当 ¥{field_work_total:,}"
        elif field_work_days > 0:
            field_work_str = f"{field_work_days}日"
            field_work_note = ''
        else:
            field_work_str = ''
            field_work_note = ''

        totals = [
            '集計', '',
            '', '',                                            # 出勤・退勤（空欄）
            '',                                                # 休憩（空欄）
            '',                                                # 実働（空欄）
            '',                                                # 所定（空欄）
            monthly['total_work'],                             # 勤務合計
            monthly['total_ot_in'],                            # 延長合計
            monthly['total_ot_out'],                           # 残業合計
            monthly['total_absence'],                          # 遅刻早退合計
            '',                                                # 欠勤（空欄）
            f"{monthly['paid_days']}日",                       # 有給日数
            monthly['paid_hours'],                             # 有給時間合計
            field_work_str,                                    # 外勤日数
            f"出勤{monthly['attend_days']}日 {field_work_note}".strip(),  # 備考
        ]

        for col_idx, val in enumerate(totals, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

    def _format_time(self, hours_float):
        """float時間 → "HH:MM" 形式の文字列に変換"""
        if hours_float is None:
            return ''
        h = int(hours_float)
        m = int((hours_float - h) * 60)
        return f"{h}:{m:02d}"
