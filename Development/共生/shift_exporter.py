#!/usr/bin/env python3
"""
shift_exporter.py — 共生（はあとふる）シフト自動生成エンジン v1.1
Pipeline_OPS-C Phase 1

使用技術: Google OR-Tools CP-SAT Solver + OpenPyXL
出力先: 20265月シフト.xlsx の「入力フォーム」「予定2対応」ワークシート

ノーススター経営サポート — 「手抜きをしない。網羅する。魂を込める。」
"""

import json
import sys
import os
from datetime import date
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple, Set
from pathlib import Path
from ortools.sat.python import cp_model
import openpyxl

# ════════════════════════════════════════════════════════════════
# §1. 定数定義
# ════════════════════════════════════════════════════════════════

# シフト記号インデックス（CP-SATの変数で使用）
S_A    = 0   # A（9:00-17:00）介護
S_B    = 1   # B（7:30-15:30）介護
S_C    = 2   # C（12:30-20:30）介護
S_D    = 3   # D（8:30-16:30）介護
S_E    = 4   # E（8:30-16:30）看護/管理者/計画作成/事務
S_YORU = 5   # 夜（17:00-翌1:00）夜勤
S_AKE  = 6   # 明（1:00-9:00）夜勤とセット
S_CH1  = 7   # 調1（6:30-8:30）調理
S_CH2  = 8   # 調2（9:30-17:30）調理
S_KYUU = 9   # 休（公休）
S_KIBO = 10  # 希（希望休）
S_YUUK = 11  # 有（有給）
S_KEN  = 12  # 健（健康診断 出勤扱い）
S_KENS = 13  # 研（研修 出勤扱い）

NUM_SHIFTS = 14

SHIFT_NAMES = ['A', 'B', 'C', 'D', 'E', '夜', '明', '調1', '調2', '休', '希', '有', '健', '研']

# 出勤としてカウントするシフト
WORK_SHIFT_IDS = {S_A, S_B, S_C, S_D, S_E, S_YORU, S_AKE, S_CH1, S_CH2, S_KEN, S_KENS}

# 休みシフト
OFF_SHIFT_IDS = {S_KYUU, S_KIBO, S_YUUK}

# 介護シフト（A,B,C,D,夜,明）
CARE_SHIFT_IDS = {S_A, S_B, S_C, S_D, S_YORU, S_AKE}

# Excel出力用: シフト記号
SHIFT_TO_EXCEL = {
    S_A: 'Ａ', S_B: 'Ｂ', S_C: 'Ｃ', S_D: 'Ｄ', S_E: 'Ｅ',
    S_YORU: '夜', S_AKE: '明',
    S_CH1: '調1', S_CH2: '調2',
    S_KYUU: '休', S_KIBO: '希', S_YUUK: '有',
    S_KEN: '健', S_KENS: '研',
}

# 曜日定数（0=月 ... 6=日）
MON, TUE, WED, THU, FRI, SAT, SUN = 0, 1, 2, 3, 4, 5, 6
WEEKDAY_NAMES = ['月', '火', '水', '木', '金', '土', '日']

# 2026年5月
YEAR = 2026
MONTH = 5
NUM_DAYS = 31

# 正社員出勤日数（5月）
FULLTIME_WORK_DAYS = 22

# ════════════════════════════════════════════════════════════════
# §2. スタッフデータモデル
# ════════════════════════════════════════════════════════════════

@dataclass
class StaffMember:
    """スタッフ1名の全情報"""
    id: int                          # 内部ID（0始まり）
    name: str                        # 氏名
    employee_code: int               # 社員コード
    role: str                        # 職種
    employment_type: str             # 雇用区分
    is_fulltime: bool                # 正社員扱いか
    night_eligible: bool             # 夜勤対象か
    allowed_shifts: Set[int]         # 可能シフト
    prescribed_days: int             # 月所定出勤日数
    max_days: int                    # 月最大出勤日数
    allowed_weekdays: Optional[Set[int]] = None  # 出勤可能曜日（None=全曜日）
    is_dual_role: bool = False       # 兼任者か
    # 入力フォーム: 兼任者でも1行のみ → (行番号,) のタプル or None
    input_form_row: Optional[int] = None
    # 予定2対応: 兼任者は複数行 → [(役割名, 行番号), ...]
    yotei2_rows: List[Tuple[str, int]] = field(default_factory=list)
    notes: str = ''

# ════════════════════════════════════════════════════════════════
# §3. スタッフマスター定義（Pipeline_OPS-C_Spec.md §3.3 準拠）
# ════════════════════════════════════════════════════════════════

def build_staff_master() -> List[StaffMember]:
    """全スタッフマスターを構築する"""
    staff = []

    # --- 0: 吉玉薫（管理者 兼 介護職員）---
    staff.append(StaffMember(
        id=0, name='吉玉薫', employee_code=0,
        role='管理者/介護職員', employment_type='常勤/兼任',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_E, S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=10, max_days=14,
        is_dual_role=True,
        input_form_row=6,  # 入力フォームRow6: 管理者/吉玉薫（1行で全シフト）
        yotei2_rows=[('管理者', 6), ('介護職員', 7)],
        notes='代表取締役。管理者(E):介護(夜明)=3:2。A,B,C,D不可。夜明は月2セット目標'
    ))

    # --- 1: 飯星まゆみ ---
    staff.append(StaffMember(
        id=1, name='飯星まゆみ', employee_code=23,
        role='介護職員', employment_type='常勤/専従',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_A, S_B, S_D, S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        input_form_row=7,
        yotei2_rows=[('介護職員', 8)],
        notes='Cはどうしても足りない場合のみ'
    ))

    # --- 2: 興梠道子 ---
    staff.append(StaffMember(
        id=2, name='興梠道子', employee_code=12,
        role='介護職員', employment_type='常勤/専従',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_A, S_B, S_D, S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        input_form_row=8,
        yotei2_rows=[('介護職員', 9)],
        notes='Cはどうしても足りない場合のみ'
    ))

    # --- 3: 小椋富美（介護 兼 調理）---
    staff.append(StaffMember(
        id=3, name='小椋富美', employee_code=14,
        role='介護職員/調理職員', employment_type='常勤/兼任',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_B, S_D, S_YORU, S_AKE, S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        is_dual_role=True,
        input_form_row=9,  # 入力フォームRow9: 介護職員/小椋富美（1行で全シフト）
        yotei2_rows=[('介護職員', 10), ('調理職員', 27)],
        notes='A,CはNG。調理2のみ（指定日のみ）'
    ))

    # --- 4: 吉玉茂春 ---
    staff.append(StaffMember(
        id=4, name='吉玉茂春', employee_code=31,
        role='介護職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_KYUU},  # 基本的に休み。空白で問題なし
        prescribed_days=0, max_days=0,
        input_form_row=None,  # 入力フォームに行なし
        yotei2_rows=[('介護職員', 11)],
        notes='空白で問題なし。入力フォームに行なし'
    ))

    # --- 5: 佐藤いそみ ---
    staff.append(StaffMember(
        id=5, name='佐藤いそみ', employee_code=32,
        role='介護職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=True,
        allowed_shifts={S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=14, max_days=15,
        input_form_row=10,
        yotei2_rows=[('介護職員', 12)],
        notes='夜、明のみ。A,B,C,DはNG'
    ))

    # --- 6: 松本きくみ ---
    staff.append(StaffMember(
        id=6, name='松本きくみ', employee_code=35,
        role='介護職員', employment_type='正社員',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_A, S_B, S_C, S_D, S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        input_form_row=11,
        yotei2_rows=[('介護職員', 13)],
        notes='2026/4入社。正社員。所定22日'
    ))

    # --- 7: 甲斐裕美子（介護 兼 調理）---
    staff.append(StaffMember(
        id=7, name='甲斐裕美子', employee_code=15,
        role='介護職員/調理職員', employment_type='非常勤/兼任',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_A, S_C, S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=12, max_days=15,
        is_dual_role=True,
        input_form_row=12,  # 入力フォームRow12: 介護職員/甲斐裕美子（1行で全シフト）
        yotei2_rows=[('介護職員', 14), ('調理職員', 28)],
        notes='日勤者。A,C（介護）。調理2のみ（指定日のみ）。B,DはNG'
    ))

    # --- 8: 辰見イツ子 ---
    staff.append(StaffMember(
        id=8, name='辰見イツ子', employee_code=6,
        role='介護職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_A, S_C, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=12, max_days=15,
        input_form_row=13,
        yotei2_rows=[('介護職員', 15)],
        notes='日勤者。A,C。B,DはNG'
    ))

    # --- 9: 甲斐静幸 ---
    staff.append(StaffMember(
        id=9, name='甲斐静幸', employee_code=2,
        role='介護職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_C, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=12, max_days=15,
        input_form_row=14,
        yotei2_rows=[('介護職員', 16)],
        notes='C限定。A,B,DはNG'
    ))

    # --- 10: 渡邊哲也（計画作成担当者 兼 介護職員）---
    staff.append(StaffMember(
        id=10, name='渡邊哲也', employee_code=7,
        role='計画作成担当者/介護職員', employment_type='常勤/兼任',
        is_fulltime=True, night_eligible=True,
        allowed_shifts={S_E, S_YORU, S_AKE, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        is_dual_role=True,
        input_form_row=15,  # 入力フォームRow15: 計画作成担当者/渡邊哲也（1行で全シフト）
        yotei2_rows=[('介護職員', 17), ('計画作成担当者', 18)],
        notes='主務=計画作成(E)、副務=介護(夜明)。E月最低12日'
    ))

    # --- 11: 甲斐充代 ---
    staff.append(StaffMember(
        id=11, name='甲斐充代', employee_code=5,
        role='機能訓練指導員', employment_type='常勤/専従',
        is_fulltime=True, night_eligible=False,
        allowed_shifts={S_A, S_B, S_D, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        input_form_row=16,
        yotei2_rows=[('機能訓練指導員', 19)],
        notes='日勤者。A,B,D限定'
    ))

    # --- 12: 田上元美 ---
    staff.append(StaffMember(
        id=12, name='田上元美', employee_code=3,
        role='生活相談員', employment_type='常勤/専従',
        is_fulltime=True, night_eligible=False,
        allowed_shifts={S_A, S_B, S_D, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        input_form_row=17,
        yotei2_rows=[('生活相談員', 20)],
        notes='日勤者。A,B,D限定'
    ))

    # --- 13: 甲斐千代美 ---
    # 所定22日。平日不足分は有給自動配置（金/月で3連休形成）
    staff.append(StaffMember(
        id=13, name='甲斐千代美', employee_code=1,
        role='看護職員', employment_type='常勤/専従',
        is_fulltime=True, night_eligible=False,
        allowed_shifts={S_E, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=FULLTIME_WORK_DAYS, max_days=FULLTIME_WORK_DAYS,
        allowed_weekdays={MON, TUE, WED, THU, FRI},
        input_form_row=18,
        yotei2_rows=[('看護職員', 21)],
        notes='E限定。土日以外。平日不足分は有給（金/月で3連休）'
    ))

    # --- 14: 椎葉玲子 ---
    staff.append(StaffMember(
        id=14, name='椎葉玲子', employee_code=9,
        role='事務職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_E, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=15, max_days=15,
        allowed_weekdays={MON, TUE, WED, THU, FRI},
        input_form_row=19,
        yotei2_rows=[('事務職員', 22)],
        notes='E限定。月所定15日。土日以外'
    ))

    # --- 15: 堀カズ子 ---
    staff.append(StaffMember(
        id=15, name='堀カズ子', employee_code=-1,
        role='調理職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_CH1, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=26, max_days=27,
        allowed_weekdays={MON, TUE, WED, THU, SAT, SUN},  # 金以外
        input_form_row=20,
        yotei2_rows=[('調理職員', 23)],
        notes='調理1のみ。金以外'
    ))

    # --- 16: 町倫子 ---
    staff.append(StaffMember(
        id=16, name='町倫子', employee_code=18,
        role='調理職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_CH1, S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=12, max_days=15,
        allowed_weekdays={WED, THU, FRI},
        input_form_row=21,
        yotei2_rows=[('調理職員', 24)],
        notes='水木は調理2、金は調理1'
    ))

    # --- 17: 甲斐幸子 ---
    staff.append(StaffMember(
        id=17, name='甲斐幸子', employee_code=22,
        role='調理職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=12, max_days=15,
        allowed_weekdays={MON, TUE, FRI},
        input_form_row=22,
        yotei2_rows=[('調理職員', 25)],
        notes='調理2のみ。月火金のみ'
    ))

    # --- 18: 甲斐美菜子 ---
    staff.append(StaffMember(
        id=18, name='甲斐美菜子', employee_code=36,
        role='調理職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=4, max_days=12,
        allowed_weekdays={SUN},
        input_form_row=23,
        yotei2_rows=[('調理職員', 26)],
        notes='調理2のみ。日のみ。社員コード36（新規発番）'
    ))

    # --- 19: 高本みち子 ---
    staff.append(StaffMember(
        id=19, name='高本みち子', employee_code=19,
        role='調理職員', employment_type='非常勤/専従',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_CH2, S_KYUU, S_KIBO, S_YUUK},
        prescribed_days=4, max_days=12,
        allowed_weekdays={SAT},
        input_form_row=24,
        yotei2_rows=[('調理職員', 29)],
        notes='調理2のみ。土のみ'
    ))

    # --- 20: 花田美幸（臨時）---
    staff.append(StaffMember(
        id=20, name='花田美幸', employee_code=33,
        role='調理職員', employment_type='臨時',
        is_fulltime=False, night_eligible=False,
        allowed_shifts={S_KYUU},
        prescribed_days=0, max_days=0,
        input_form_row=25,
        yotei2_rows=[('調理職員', 30)],
        notes='臨時。基本はシフトに組まない'
    ))

    return staff


# ════════════════════════════════════════════════════════════════
# §4. 日付ユーティリティ
# ════════════════════════════════════════════════════════════════

def get_weekday(day_index: int) -> int:
    """day_index（0=5/1）の曜日を返す（0=月 ... 6=日）"""
    return date(YEAR, MONTH, day_index + 1).weekday()

def is_weekend(day_index: int) -> bool:
    """土日かどうか"""
    wd = get_weekday(day_index)
    return wd == SAT or wd == SUN

def is_sunday(day_index: int) -> bool:
    """日曜かどうか"""
    return get_weekday(day_index) == SUN


# ════════════════════════════════════════════════════════════════
# §5. 希望休データ読込
# ════════════════════════════════════════════════════════════════

def load_requests(filepath: str) -> Dict[str, Dict]:
    """
    希望休・有給データをJSONから読み込む。
    形式: {"飯星まゆみ": {"希望休": [5, 12, 25], "有給": [15]}, ...}
    日付は月内の日（1〜31）。
    """
    if not os.path.exists(filepath):
        print(f"  [INFO] 希望休ファイルなし。デフォルトで生成します。")
        return {}
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"  [OK] 希望休ファイル読込完了: {len(data)}名分")
    return data


# ════════════════════════════════════════════════════════════════
# §6. Excel行・列マッピング
# ════════════════════════════════════════════════════════════════

# 入力フォーム: 5/1=G列(col7), 5/2=H列(col8) ... 5/31=AK列(col37)
INPUT_FORM_COL_OFFSET = 7  # day_index=0 → col 7

# 予定2対応: 5/1=E列(col5), 5/2=F列(col6) ... 5/31=AI列(col35)
YOTEI2_COL_OFFSET = 5  # day_index=0 → col 5


# ════════════════════════════════════════════════════════════════
# §7. CP-SAT ソルバ
# ════════════════════════════════════════════════════════════════

class ShiftSolver:
    """全17ルールを制約充足問題として定式化し、OR-Tools CP-SATで求解する。"""

    def __init__(self, staff: List[StaffMember], requests: Dict[str, Dict]):
        self.staff = staff
        self.requests = requests
        self.num_staff = len(staff)
        self.num_days = NUM_DAYS
        self.model = cp_model.CpModel()
        self.solver = cp_model.CpSolver()

        # x[p][d][s] = 1 iff スタッフpが日dにシフトsを担当
        self.x = {}
        for p in range(self.num_staff):
            self.x[p] = {}
            for d in range(self.num_days):
                self.x[p][d] = {}
                for s in range(NUM_SHIFTS):
                    self.x[p][d][s] = self.model.NewBoolVar(f'x_{p}_{d}_{s}')

        # 注: is_working補助変数は使用しない（AddMaxEqualityの同時定義で矛盾するため）
        # 出勤日数は _work_sum() ヘルパーで直接計算する

        # 介護職員ID（吉玉茂春除外）
        self.care_staff_ids = [st.id for st in self.staff
                               if '介護職員' in st.role and st.id != 4]

        # 夜勤対象ID
        self.night_staff_ids = [st.id for st in self.staff if st.night_eligible]

        # 機能訓練/生活相談ID
        self.ft_sw_ids = [11, 12]

        self.solution = None

    def build_all_constraints(self):
        """全制約を構築する"""
        print("[SOLVER] 制約構築開始...")

        self._constraint_basic()
        print("  [1/17] 基本制約（1日1シフト + 出勤補助変数）")

        self._constraint_rule11_allowed_shifts()
        print("  [2/17] ルール11: 個人別可能/不可シフト + 曜日制限")

        self._constraint_rule14_requests()
        print("  [3/17] ルール14: 希望休・有給厳守")

        self._constraint_rule1_fulltime_days()
        print("  [4/17] ルール1: 正社員出勤日数")

        self._constraint_rule2_parttime_days()
        print("  [5/17] ルール2: パート出勤日数")

        self._constraint_rule6_night_dawn_pair()
        print("  [6/17] ルール6: 夜明同一職員連続")

        self._constraint_rule4_care_24h()
        print("  [7/17] ルール4: 介護24時間配置")

        self._constraint_rule5_care_base_pattern()
        print("  [8/17] ルール5: 介護基本配置パターン")

        self._constraint_rule8_one_per_shift()
        print("  [9/17] ルール8: 各シフト1名配置")

        self._constraint_rule9_ft_sw_weekend()
        print("  [10/17] ルール9: 機能訓練/生活相談 土日制約")

        self._constraint_rule12_dual_ratio()
        print("  [11/17] ルール12: 兼任比率")

        self._constraint_rule13_cooking()
        print("  [12/17] ルール13: 調理配置")

        self._constraint_rule15_night_rhythm()
        print("  [13/17] ルール15: 夜明リズム")

        self._constraint_rule17_rest_interval()
        print("  [14/17] ルール17: 休日間隔")

        self._set_objective()
        print("  [15/17] 目的関数（Soft制約）設定完了")

        print("[SOLVER] 制約構築完了。")

    # ─── ヘルパー: 出勤日数の合計変数リスト ───
    def _work_sum(self, p: int) -> list:
        """スタッフpの全日の出勤シフト変数をフラットリストで返す"""
        result = []
        st = self.staff[p]
        for d in range(self.num_days):
            for s in WORK_SHIFT_IDS:
                if s in st.allowed_shifts:
                    result.append(self.x[p][d][s])
        return result

    def _work_on_day(self, p: int, d: int) -> list:
        """スタッフpが日dに出勤する全シフト変数のリスト"""
        st = self.staff[p]
        return [self.x[p][d][s] for s in WORK_SHIFT_IDS if s in st.allowed_shifts]

    # ─── 基本制約 ───
    def _constraint_basic(self):
        for p in range(self.num_staff):
            for d in range(self.num_days):
                # 1日1シフト
                self.model.AddExactlyOne(
                    [self.x[p][d][s] for s in range(NUM_SHIFTS)]
                )

    # ─── ルール11: 可能/不可シフト + 曜日制限 ───
    def _constraint_rule11_allowed_shifts(self):
        for st in self.staff:
            p = st.id
            for d in range(self.num_days):
                wd = get_weekday(d)
                for s in range(NUM_SHIFTS):
                    if s not in st.allowed_shifts:
                        self.model.Add(self.x[p][d][s] == 0)
                    elif st.allowed_weekdays is not None and wd not in st.allowed_weekdays:
                        if s in WORK_SHIFT_IDS:
                            self.model.Add(self.x[p][d][s] == 0)

    # ─── ルール14: 希望休・有給厳守 ───
    def _constraint_rule14_requests(self):
        for st in self.staff:
            p = st.id
            if st.name in self.requests:
                req = self.requests[st.name]
                if '希望休' in req:
                    for day_num in req['希望休']:
                        d = day_num - 1
                        if 0 <= d < self.num_days and S_KIBO in st.allowed_shifts:
                            self.model.Add(self.x[p][d][S_KIBO] == 1)
                if '有給' in req:
                    for day_num in req['有給']:
                        d = day_num - 1
                        if 0 <= d < self.num_days and S_YUUK in st.allowed_shifts:
                            self.model.Add(self.x[p][d][S_YUUK] == 1)

    # ─── ルール1: 正社員出勤日数 ───
    def _constraint_rule1_fulltime_days(self):
        for st in self.staff:
            if not st.is_fulltime or st.id == 4:
                continue
            p = st.id
            total_work = self._work_sum(p)

            if st.id == 0:  # 吉玉薫: 目標10日（8〜14で柔軟に）
                self.model.Add(sum(total_work) >= 8)
                self.model.Add(sum(total_work) <= 14)
            elif st.id == 13:  # 甲斐千代美: 平日不足分は有給自動配置
                self._constraint_kaichiyomi_paid_leave()
            else:
                # 松本きくみも含め、全正社員は所定FULLTIME_WORK_DAYS日
                self.model.Add(sum(total_work) == FULLTIME_WORK_DAYS)

    # ─── 甲斐千代美: 有給自動配置ロジック ───
    def _constraint_kaichiyomi_paid_leave(self):
        """
        甲斐千代美（id=13）は正社員と同じ所定日数（出勤+有給=FULLTIME_WORK_DAYS）。
        Eシフトは平日のみだが、土日は「制度上の休み」ではなく
        「正社員と同じ休日数を土日に充てている」だけ。

        ケースA（平日 < 所定）:
          平日だけでは所定に届かないため、不足分を土日の任意の日に有給配置。
          例: 5月 平日21 < 所定22 → 土日1日を有給

        ケースB（平日 > 所定）:
          平日に余裕があるため、余剰分を金曜or月曜に有給配置して3連休形成。
          例: 7月 平日23 > 所定22 → 金/月1日を有給

        ケースC（平日 == 所定）:
          有給なし。全平日出勤。
        """
        p = 13
        st = self.staff[p]

        # 平日の日数・金曜/月曜/土日のday_indexを収集
        weekday_count = 0
        mon_days = []
        fri_days = []
        weekend_days = []
        for d in range(self.num_days):
            wd = get_weekday(d)
            if wd in st.allowed_weekdays:
                weekday_count += 1
            if wd == MON:
                mon_days.append(d)
            if wd == FRI:
                fri_days.append(d)
            if wd == SAT or wd == SUN:
                weekend_days.append(d)

        total_work = self._work_sum(p)

        if weekday_count == FULLTIME_WORK_DAYS:
            # ケースC: 平日ちょうど所定日数。有給なし
            self.model.Add(sum(total_work) == FULLTIME_WORK_DAYS)
            print(f"    [INFO] 甲斐千代美: 平日{weekday_count}日 == 所定{FULLTIME_WORK_DAYS}日 → 有給なし")

        elif weekday_count < FULLTIME_WORK_DAYS:
            # ケースA: 平日不足 → 不足分を土日に有給配置
            shortage = FULLTIME_WORK_DAYS - weekday_count
            # 実出勤 = 全平日
            self.model.Add(sum(total_work) == weekday_count)
            # 有給 = 不足分（土日に配置）
            yuuk_vars = [self.x[p][d][S_YUUK] for d in range(self.num_days)]
            self.model.Add(sum(yuuk_vars) == shortage)
            # 有給は土日のみに配置
            weekend_set = set(weekend_days)
            for d in range(self.num_days):
                if d not in weekend_set:
                    self.model.Add(self.x[p][d][S_YUUK] == 0)
            print(f"    [INFO] 甲斐千代美: 平日{weekday_count}日 < 所定{FULLTIME_WORK_DAYS}日 → "
                  f"実出勤{weekday_count}日 + 有給{shortage}日（土日に配置）")

        else:
            # ケースB: 平日余剰 → 余剰分を金/月に有給配置して3連休
            excess = weekday_count - FULLTIME_WORK_DAYS
            actual_work = FULLTIME_WORK_DAYS - excess
            # 実出勤 = 所定 - 有給分
            self.model.Add(sum(total_work) == actual_work)
            # 有給 = 余剰分（金/月に配置）
            yuuk_vars = [self.x[p][d][S_YUUK] for d in range(self.num_days)]
            self.model.Add(sum(yuuk_vars) == excess)
            # 有給は金曜or月曜のみに配置（3連休形成）
            valid_yuuk_days = set(mon_days + fri_days)
            for d in range(self.num_days):
                if d not in valid_yuuk_days:
                    self.model.Add(self.x[p][d][S_YUUK] == 0)
            print(f"    [INFO] 甲斐千代美: 平日{weekday_count}日 > 所定{FULLTIME_WORK_DAYS}日 → "
                  f"実出勤{actual_work}日 + 有給{excess}日（金/月に配置で3連休）")

    # ─── ルール2: パート出勤日数 ───
    def _constraint_rule2_parttime_days(self):
        for st in self.staff:
            if st.is_fulltime or (st.prescribed_days == 0 and st.max_days == 0):
                continue
            p = st.id
            total_work = self._work_sum(p)
            self.model.Add(sum(total_work) >= st.prescribed_days)
            self.model.Add(sum(total_work) <= st.max_days)

    # ─── ルール6: 夜明同一職員連続 ───
    def _constraint_rule6_night_dawn_pair(self):
        for p in self.night_staff_ids:
            for d in range(self.num_days - 1):
                # 夜(d) → 翌日明(d+1)
                self.model.Add(self.x[p][d + 1][S_AKE] >= self.x[p][d][S_YORU])
            for d in range(1, self.num_days):
                # 明(d) → 前日夜(d-1)
                self.model.Add(self.x[p][d - 1][S_YORU] >= self.x[p][d][S_AKE])

    # ─── ルール4: 介護24時間配置 ───
    def _constraint_rule4_care_24h(self):
        for d in range(self.num_days):
            # 夜勤: 毎日ちょうど1名
            night_vars = [self.x[p][d][S_YORU] for p in self.night_staff_ids]
            self.model.Add(sum(night_vars) == 1)
            # 明け: 毎日ちょうど1名
            dawn_vars = [self.x[p][d][S_AKE] for p in self.night_staff_ids]
            self.model.Add(sum(dawn_vars) == 1)

    # ─── ルール5: 介護基本配置パターン ───
    def _constraint_rule5_care_base_pattern(self):
        for d in range(self.num_days):
            # 介護職員からCシフト: 毎日ちょうど1名
            c_vars = [self.x[p][d][S_C] for p in self.care_staff_ids
                      if S_C in self.staff[p].allowed_shifts]
            if c_vars:
                self.model.Add(sum(c_vars) == 1)

            # 介護職員からB or Dシフト: 毎日最大1名
            bd_vars = []
            for p in self.care_staff_ids:
                for s in [S_B, S_D]:
                    if s in self.staff[p].allowed_shifts:
                        bd_vars.append(self.x[p][d][s])
            if bd_vars:
                self.model.Add(sum(bd_vars) <= 1)

    # ─── ルール8: 各シフト1名配置 ───
    def _constraint_rule8_one_per_shift(self):
        target_ids = list(self.care_staff_ids) + self.ft_sw_ids

        for d in range(self.num_days):
            for s in [S_A, S_B, S_D]:
                vars_for_shift = [self.x[p][d][s] for p in target_ids
                                  if s in self.staff[p].allowed_shifts]
                if not vars_for_shift:
                    continue
                if s == S_A and is_sunday(d):
                    self.model.Add(sum(vars_for_shift) <= 1)
                else:
                    self.model.Add(sum(vars_for_shift) == 1)

    # ─── ルール9: 機能訓練/生活相談 土日制約 ───
    def _constraint_rule9_ft_sw_weekend(self):
        for p_id in self.ft_sw_ids:
            for d in range(self.num_days - 1):
                if get_weekday(d) == SAT:
                    d_sun = d + 1
                    if d_sun >= self.num_days:
                        break
                    # 土曜B/D + 日曜B/D の同時禁止
                    for s1 in [S_B, S_D]:
                        for s2 in [S_B, S_D]:
                            if (s1 in self.staff[p_id].allowed_shifts and
                                    s2 in self.staff[p_id].allowed_shifts):
                                self.model.AddBoolOr([
                                    self.x[p_id][d][s1].Not(),
                                    self.x[p_id][d_sun][s2].Not()
                                ])

    # ─── ルール12: 兼任比率 ───
    def _constraint_rule12_dual_ratio(self):
        # 吉玉薫: E >= 夜明の1.5倍（3:2比率）
        p = 0
        e_days = [self.x[p][d][S_E] for d in range(self.num_days)]
        night_days = [self.x[p][d][S_YORU] for d in range(self.num_days)]
        self.model.Add(sum(e_days) * 2 >= sum(night_days) * 3)

        # 渡邊哲也: E >= 12日、E > 夜明
        p = 10
        e_days_w = [self.x[p][d][S_E] for d in range(self.num_days)]
        night_days_w = [self.x[p][d][S_YORU] for d in range(self.num_days)]
        self.model.Add(sum(e_days_w) >= 12)
        self.model.Add(sum(e_days_w) * 2 >= sum(night_days_w) * 3)

    # ─── ルール13: 調理配置 ───
    def _constraint_rule13_cooking(self):
        """調理1・調理2を毎日1名ずつ配置する"""
        # 調理1可能: 堀(15), 町(16/金のみ)
        # 調理2可能: 町(16/水木), 甲斐幸子(17), 甲斐美菜子(18), 高本(19), 小椋(3), 甲斐裕美子(7)
        cook1_ids = [15, 16]
        cook2_ids = [16, 17, 18, 19, 3, 7]

        for d in range(self.num_days):
            wd = get_weekday(d)

            # 調理1
            ch1_vars = []
            for p in cook1_ids:
                if S_CH1 in self.staff[p].allowed_shifts:
                    if self.staff[p].allowed_weekdays is None or wd in self.staff[p].allowed_weekdays:
                        ch1_vars.append(self.x[p][d][S_CH1])
            if ch1_vars:
                self.model.Add(sum(ch1_vars) == 1)

            # 調理2
            ch2_vars = []
            for p in cook2_ids:
                if S_CH2 in self.staff[p].allowed_shifts:
                    if self.staff[p].allowed_weekdays is None or wd in self.staff[p].allowed_weekdays:
                        ch2_vars.append(self.x[p][d][S_CH2])
            if ch2_vars:
                self.model.Add(sum(ch2_vars) == 1)

        # 町倫子の曜日別固定: 水木→調理2、金→調理1
        p_machi = 16
        for d in range(self.num_days):
            wd = get_weekday(d)
            if wd in {WED, THU}:
                self.model.Add(self.x[p_machi][d][S_CH2] == 1)
            elif wd == FRI:
                self.model.Add(self.x[p_machi][d][S_CH1] == 1)

    # ─── ルール15: 夜明リズム ───
    def _constraint_rule15_night_rhythm(self):
        for p in self.night_staff_ids:
            for d in range(self.num_days):
                # 夜勤の前日は出勤不可（明け除く）
                if d >= 1:
                    for s in WORK_SHIFT_IDS:
                        if s == S_AKE:
                            continue
                        if s in self.staff[p].allowed_shifts:
                            self.model.AddBoolOr([
                                self.x[p][d][S_YORU].Not(),
                                self.x[p][d - 1][s].Not()
                            ])

                # 明け翌日にB,C,D,夜は不可（Aは可、休みは可）
                if d < self.num_days - 1:
                    for s in [S_B, S_C, S_D, S_YORU]:
                        if s in self.staff[p].allowed_shifts:
                            self.model.AddBoolOr([
                                self.x[p][d][S_AKE].Not(),
                                self.x[p][d + 1][s].Not()
                            ])

    # ─── ルール17: 休日間隔 ───
    def _constraint_rule17_rest_interval(self):
        """連続出勤は最大6日まで（7連勤禁止）"""
        # 堀カズ子(15)は金以外26日/26日出勤で連勤不可避のため除外
        excluded_ids = {4, 15, 20}  # 吉玉茂春, 堀カズ子, 花田
        for st in self.staff:
            if st.id in excluded_ids or st.prescribed_days == 0:
                continue
            p = st.id
            for d in range(self.num_days - 6):
                work_7 = []
                for dd in range(d, d + 7):
                    work_7.extend(self._work_on_day(p, dd))
                self.model.Add(sum(work_7) <= 6)

    # ─── 目的関数（Soft制約） ───
    def _set_objective(self):
        obj_terms = []

        # パターン①②優先: 介護職員からのB or D配置に報酬
        for d in range(self.num_days):
            for p in self.care_staff_ids:
                if S_B in self.staff[p].allowed_shifts:
                    obj_terms.append(self.x[p][d][S_B] * 10)
                if S_D in self.staff[p].allowed_shifts:
                    obj_terms.append(self.x[p][d][S_D] * 10)

        # 吉玉薫: 出勤日数を10日にできるだけ近づける
        p0_work = self._work_sum(0)
        dev_above = self.model.NewIntVar(0, 14, 'p0_dev_above')
        dev_below = self.model.NewIntVar(0, 14, 'p0_dev_below')
        self.model.Add(sum(p0_work) - 10 == dev_above - dev_below)
        obj_terms.append(dev_above * (-2))
        obj_terms.append(dev_below * (-2))

        if obj_terms:
            self.model.Maximize(sum(obj_terms))

    # ─── 求解 ───
    def solve(self, time_limit_seconds: int = 120) -> bool:
        print(f"\n[SOLVER] 求解開始（制限時間: {time_limit_seconds}秒）...")
        self.solver.parameters.max_time_in_seconds = time_limit_seconds
        self.solver.parameters.num_workers = 8

        status = self.solver.Solve(self.model)

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            label = "最適解" if status == cp_model.OPTIMAL else "実行可能解"
            print(f"[SOLVER] {label}を発見。目的関数値: {self.solver.ObjectiveValue()}")
            self._extract_solution()
            return True
        elif status == cp_model.INFEASIBLE:
            print("[SOLVER] *** 解なし（制約が矛盾）。制約を緩和してください。 ***")
            return False
        else:
            print(f"[SOLVER] ステータス: {status}")
            return False

    def _extract_solution(self):
        self.solution = {}
        for p in range(self.num_staff):
            self.solution[p] = {}
            for d in range(self.num_days):
                for s in range(NUM_SHIFTS):
                    if self.solver.Value(self.x[p][d][s]) == 1:
                        self.solution[p][d] = s
                        break

    def get_shift(self, staff_id: int, day_index: int) -> int:
        if self.solution is None:
            raise ValueError("まだ求解されていません")
        return self.solution[staff_id][day_index]


# ════════════════════════════════════════════════════════════════
# §8. Excel書込みエンジン
# ════════════════════════════════════════════════════════════════

class ExcelExporter:
    """シフト結果を20265月シフト.xlsxに書き込む"""

    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path
        self.wb = openpyxl.load_workbook(template_path)

    def write_shifts(self, solver: ShiftSolver, staff: List[StaffMember]):
        """入力フォームと予定2対応の両シートにシフトを書き込む"""
        self._write_input_form(solver, staff)
        self._write_yotei2(solver, staff)
        self.wb.save(self.output_path)
        print(f"\n[EXCEL] 保存完了: {self.output_path}")

    def _write_input_form(self, solver: ShiftSolver, staff: List[StaffMember]):
        """
        入力フォーム: 兼任者も1行にまとめて全シフトをそのまま書く。
        """
        ws = self.wb['入力フォーム']
        print("\n[EXCEL] 入力フォームに書込み中...")

        for st in staff:
            if st.input_form_row is None:
                continue  # 吉玉茂春: 入力フォームに行なし

            row = st.input_form_row
            for d in range(NUM_DAYS):
                shift_id = solver.get_shift(st.id, d)
                col = INPUT_FORM_COL_OFFSET + d
                ws.cell(row=row, column=col, value=SHIFT_TO_EXCEL.get(shift_id, ''))

        print(f"  [OK] 入力フォーム書込み完了")

    def _write_yotei2(self, solver: ShiftSolver, staff: List[StaffMember]):
        """
        予定2対応: 兼任者は複数行。ロールに対応するシフトのみ書き込む。
        """
        ws = self.wb['予定2対応']
        print("[EXCEL] 予定2対応に書込み中...")

        for st in staff:
            for role_name, row in st.yotei2_rows:
                for d in range(NUM_DAYS):
                    shift_id = solver.get_shift(st.id, d)
                    col = YOTEI2_COL_OFFSET + d
                    val = self._get_yotei2_value(st, role_name, shift_id)
                    ws.cell(row=row, column=col, value=val)

        print(f"  [OK] 予定2対応書込み完了")

    def _get_yotei2_value(self, st: StaffMember, role_name: str,
                          shift_id: int) -> str:
        """
        予定2対応の兼任者用: そのロールに該当するシフトのみ記号を出力。
        該当しない場合は「休」を出力。
        """
        if not st.is_dual_role:
            return SHIFT_TO_EXCEL.get(shift_id, '')

        # 休み系はどのロール行にもそのまま出力
        if shift_id in OFF_SHIFT_IDS:
            return SHIFT_TO_EXCEL[shift_id]

        # 吉玉薫（管理者/介護職員）
        if st.id == 0:
            if role_name == '管理者':
                return SHIFT_TO_EXCEL[shift_id] if shift_id == S_E else SHIFT_TO_EXCEL[S_KYUU]
            elif role_name == '介護職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id in {S_YORU, S_AKE} else SHIFT_TO_EXCEL[S_KYUU]

        # 渡邊哲也（計画作成担当者/介護職員）
        elif st.id == 10:
            if role_name == '計画作成担当者':
                return SHIFT_TO_EXCEL[shift_id] if shift_id == S_E else SHIFT_TO_EXCEL[S_KYUU]
            elif role_name == '介護職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id in {S_YORU, S_AKE} else SHIFT_TO_EXCEL[S_KYUU]

        # 小椋富美（介護職員/調理職員）
        elif st.id == 3:
            if role_name == '介護職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id in {S_B, S_D, S_YORU, S_AKE} else SHIFT_TO_EXCEL[S_KYUU]
            elif role_name == '調理職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id == S_CH2 else SHIFT_TO_EXCEL[S_KYUU]

        # 甲斐裕美子（介護職員/調理職員）
        elif st.id == 7:
            if role_name == '介護職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id in {S_A, S_C} else SHIFT_TO_EXCEL[S_KYUU]
            elif role_name == '調理職員':
                return SHIFT_TO_EXCEL[shift_id] if shift_id == S_CH2 else SHIFT_TO_EXCEL[S_KYUU]

        return SHIFT_TO_EXCEL.get(shift_id, '')


# ════════════════════════════════════════════════════════════════
# §9. コンソール表示
# ════════════════════════════════════════════════════════════════

def print_shift_table(solver: ShiftSolver, staff: List[StaffMember]):
    """シフト表をコンソールに表示する"""
    print("\n" + "=" * 100)
    print("  2026年5月 シフト表（自動生成結果）")
    print("=" * 100)

    # ヘッダー
    header = f"{'名前':<10}"
    for d in range(NUM_DAYS):
        header += f"{d+1:>3}"
    header += " |出勤"
    print(header)

    wd_line = f"{'':10}"
    for d in range(NUM_DAYS):
        wd_line += f"  {WEEKDAY_NAMES[get_weekday(d)]}"
    print(wd_line)
    print("-" * 100)

    for st in staff:
        if st.prescribed_days == 0 and st.max_days == 0:
            continue

        line = f"{st.name:<10}"
        work_count = 0
        for d in range(NUM_DAYS):
            shift_id = solver.get_shift(st.id, d)
            sn = SHIFT_NAMES[shift_id]
            if shift_id in WORK_SHIFT_IDS:
                work_count += 1
            line += f"{sn:>3}"
        line += f" |{work_count:3d}"
        print(line)

    print("=" * 100)

    # 日別 介護配置サマリー
    print("\n[日別 介護配置]")
    for d in range(NUM_DAYS):
        wd = WEEKDAY_NAMES[get_weekday(d)]
        assignments = {}
        for st in staff:
            sid = solver.get_shift(st.id, d)
            sn = SHIFT_NAMES[sid]
            if sn in ('A', 'B', 'C', 'D', '夜', '明'):
                assignments[sn] = st.name
        parts = []
        for key in ['B', 'C', 'D', 'A', '夜', '明']:
            if key in assignments:
                parts.append(f"{key}={assignments[key]}")
        print(f"  5/{d+1:2d}({wd}): {', '.join(parts)}")


# ════════════════════════════════════════════════════════════════
# §10. メイン実行
# ════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  共生（はあとふる）シフト自動生成エンジン v1.1")
    print("  Pipeline_OPS-C Phase 1")
    print("=" * 60)

    base_dir = Path(__file__).parent
    template_path = base_dir / '20265月シフト.xlsx'
    output_path = base_dir / '20265月シフト_generated.xlsx'
    requests_path = base_dir / 'kiboukyuu.json'

    if not template_path.exists():
        print(f"[ERROR] テンプレート不在: {template_path}")
        sys.exit(1)

    print("\n[1/5] スタッフマスター構築...")
    staff = build_staff_master()
    print(f"  登録: {len(staff)}名")

    print("\n[2/5] 希望休データ読込...")
    requests = load_requests(str(requests_path))
    if requests:
        for name, req in requests.items():
            print(f"  {name}: {req}")

    print("\n[3/5] CP-SATソルバ構築...")
    solver = ShiftSolver(staff, requests)
    solver.build_all_constraints()

    print("\n[4/5] シフト最適化...")
    success = solver.solve(time_limit_seconds=120)

    if not success:
        print("\n[FAILED] シフト生成失敗。制約を確認してください。")
        sys.exit(1)

    print_shift_table(solver, staff)

    print("\n[5/5] Excel書込み...")
    exporter = ExcelExporter(str(template_path), str(output_path))
    exporter.write_shifts(solver, staff)

    print("\n" + "=" * 60)
    print("  完了")
    print(f"  出力: {output_path}")
    print("=" * 60)


if __name__ == '__main__':
    main()
