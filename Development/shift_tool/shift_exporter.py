"""
AIシフト自動作成ツール - Excel出力モジュール v2
既存の入力フォーム（結果2シート）の構造を忠実に再現する。
"""

import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional

from shift_solver import (
    CARE_SHIFTS, COOK_SHIFTS, ALL_SHIFTS, OFF_TYPES,
    WORK_COUNTED, WEEKDAY_JP, get_weekday_jp, is_weekend, is_sunday
)


# ============================================================
# スタイル定義
# ============================================================

COLORS = {
    'header_bg': 'D9E2F3',
    'weekend_bg': 'FFF2CC',
    'night_font': '0000FF',
    'dawn_font': '0000FF',
    'off_font': '808080',
    'hope_font': 'FF6600',
    'paid_font': '009900',
    'violation_bg': 'FFD9D9',
    'success_bg': 'D9FFD9',
    'overtime_bg': 'FFFFCC',
    'summary_bg': 'E2EFDA',
    'cook_bg': 'FCE4EC',
    'admin_bg': 'E8EAF6',
}

SHIFT_DISPLAY = {
    'A': 'Ａ', 'B': 'Ｂ', 'C': 'Ｃ', 'D': 'Ｄ', 'E': 'Ｅ',
    'night': '夜', 'dawn': '明',
    'cook1': '調1', 'cook2': '調2',
    '休': '休', '希': '希', '有': '有', '研': '研', '健': '健', 'off': '休',
}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def get_shift_display(code: str) -> str:
    return SHIFT_DISPLAY.get(code, code)


def get_cell_font(code: str) -> Font:
    if code in ['night', '夜']:
        return Font(color=COLORS['night_font'], bold=True, size=10)
    elif code in ['dawn', '明']:
        return Font(color=COLORS['dawn_font'], bold=True, size=10)
    elif code in OFF_TYPES + ['off', '休']:
        return Font(color=COLORS['off_font'], size=10)
    elif code == '希':
        return Font(color=COLORS['hope_font'], size=10)
    elif code == '有':
        return Font(color=COLORS['paid_font'], bold=True, size=10)
    return Font(size=10)


# ============================================================
# スタッフ行の順序定義（入力フォームに合わせる）
# ============================================================

# 入力フォームでは兼任者が複数行に分かれている
# 順序: 管理者(吉玉), 介護(吉玉), 介護(飯星), 介護(興梠), 介護(小椋),
#       介護(吉玉茂春), 介護(佐藤), 介護(芹口), 介護(甲斐裕美子),
#       介護(辰見), 介護(甲斐静幸), 介護(渡邊),
#       計画作成(渡邊), 機能訓練(甲斐充代), 生活相談(田上),
#       看護(甲斐千代美), 事務(椎葉),
#       調理(堀), 調理(町), 調理(甲斐幸子), 調理(甲斐嘉子),
#       調理(小椋), 調理(甲斐裕美子), 調理(高本), 調理(花田)


def build_staff_rows(result: dict) -> list:
    """入力フォームの行順序に合わせてスタッフ行を構築する。
    兼任者は同じスケジュールを複数行に表示する。"""
    staff_by_name = {}
    for sr in result['staff']:
        staff_by_name[sr['name']] = sr

    # 入力フォームの行定義（役職, 雇用形態, 氏名）
    row_defs = [
        ("管理者", "常勤/兼任", "吉玉 薫"),
        ("介護職員", "常勤/兼任", "吉玉 薫"),
        ("介護職員", "常勤/専従", "飯星 まゆみ"),
        ("介護職員", "常勤/専従", "興梠 道子"),
        ("介護職員", "常勤/兼任", "小椋 富美"),
        ("介護職員", "非常勤/専従", "吉玉 茂春"),
        ("介護職員", "非常勤/専従", "佐藤 いそみ"),
        ("介護職員", "非常勤/専従", "芹口 菅子"),
        ("介護職員", "非常勤/兼任", "甲斐 裕美子"),
        ("介護職員", "非常勤/専従", "辰見 イツ子"),
        ("介護職員", "非常勤/専従", "甲斐 静幸"),
        ("介護職員", "常勤/兼任", "渡邊 哲也"),
        ("計画作成担当者", "常勤/兼任", "渡邊 哲也"),
        ("機能訓練指導員", "常勤/専従", "甲斐 充代"),
        ("生活相談員", "常勤/専従", "田上 元美"),
        ("看護職員", "常勤/専従", "甲斐 千代美"),
        ("事務職員", "非常勤/専従", "椎葉 玲子"),
        ("調理職員", "常勤/専従", "堀 カズ子"),
        ("調理職員", "常勤/専従", "町 倫子"),
        ("調理職員", "常勤/専従", "甲斐 幸子"),
        ("調理職員", "常勤/専従", "甲斐 嘉子"),
        ("調理職員", "常勤/兼任", "小椋 富美"),
        ("調理職員", "非常勤/兼任", "甲斐 裕美子"),
        ("調理職員", "常勤/専従", "高本 みち子"),
        ("調理職員", "常勤/専従", "花田 友代"),
    ]

    rows = []
    for role, emp, name in row_defs:
        sr = staff_by_name.get(name)
        if sr:
            rows.append({
                'role': role,
                'employment': emp,
                'name': name,
                'schedule': sr['schedule'],
                'summary': sr['summary'],
                'display_name': name.replace(' ', '　'),  # 全角スペース
            })
        else:
            # スタッフが見つからない場合は空行
            rows.append({
                'role': role,
                'employment': emp,
                'name': name,
                'schedule': {},
                'summary': {'work_days': 0, 'paid_leave': 0, 'public_holidays': 0},
                'display_name': name.replace(' ', '　'),
            })

    return rows


def export_to_excel(result: dict, output_path: str):
    """シフト結果をExcelファイルに出力する（入力フォーム形式）"""
    wb = Workbook()

    year = result['year']
    month = result['month']
    num_days = result['num_days']
    facility_name = result['facility_name']

    # ============================================================
    # 結果シート（入力フォーム形式）
    # ============================================================
    ws = wb.active
    ws.title = "結果"

    # --- Row 2: タイトル ---
    ws.cell(row=2, column=2).value = "【結果】"
    ws.cell(row=2, column=2).font = Font(bold=True, size=12)
    ws.cell(row=2, column=4).value = "シフト表"
    ws.cell(row=2, column=4).font = Font(bold=True, size=12)
    ws.cell(row=2, column=5).value = datetime.date(year, month, 1)
    ws.cell(row=2, column=5).number_format = 'YYYY/M/D'
    ws.cell(row=2, column=8).value = "〜"
    ws.cell(row=2, column=9).value = datetime.date(
        year, month, num_days)
    ws.cell(row=2, column=9).number_format = 'YYYY/M/D'

    # --- Row 3: 施設名 ---
    ws.cell(row=3, column=num_days // 2 + 4).value = facility_name
    ws.cell(row=3, column=num_days // 2 + 4).font = Font(size=10)

    # --- Row 4: 日付行 ---
    DAY_COL_START = 5  # E列から日付を開始（B=役職, C=雇用形態, D=氏名）

    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        d = datetime.date(year, month, day)
        cell = ws.cell(row=4, column=col)
        cell.value = d
        cell.number_format = 'M/D'
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if is_weekend(year, month, day):
            cell.fill = PatternFill(
                start_color=COLORS['weekend_bg'],
                end_color=COLORS['weekend_bg'], fill_type='solid')

    # --- Row 5: 曜日行 ---
    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        wd = get_weekday_jp(year, month, day)
        cell = ws.cell(row=5, column=col)
        cell.value = wd
        cell.font = Font(size=9,
                         color='FF0000' if wd in ['土', '日'] else '000000')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if is_weekend(year, month, day):
            cell.fill = PatternFill(
                start_color=COLORS['weekend_bg'],
                end_color=COLORS['weekend_bg'], fill_type='solid')

    # --- 集計列ヘッダー（Row 5） ---
    summary_col_start = DAY_COL_START + num_days + 1  # 1列空けて集計
    summary_headers = ['Ａ', 'Ｂ', 'Ｄ', 'Ｅ', 'C', '夜', '明',
                       '他', '調1', '調2', '有', '休', '出勤計',
                       '出勤上限', '出勤日数チェック']

    for i, label in enumerate(summary_headers):
        cell = ws.cell(row=5, column=summary_col_start + i)
        cell.value = label
        cell.font = Font(bold=True, size=8)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(
            start_color=COLORS['header_bg'],
            end_color=COLORS['header_bg'], fill_type='solid')

    # --- 列幅設定 ---
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    for day in range(1, num_days + 1):
        cl = get_column_letter(DAY_COL_START + day - 1)
        ws.column_dimensions[cl].width = 4.5
    for i in range(len(summary_headers)):
        cl = get_column_letter(summary_col_start + i)
        ws.column_dimensions[cl].width = 5 if i < 12 else 7

    # --- スタッフ行 ---
    staff_rows = build_staff_rows(result)
    DATA_ROW_START = 6

    for idx, row_data in enumerate(staff_rows):
        row = DATA_ROW_START + idx
        role = row_data['role']
        emp = row_data['employment']
        name = row_data['display_name']
        schedule = row_data['schedule']
        summary = row_data['summary']

        # B列: 役職
        cell = ws.cell(row=row, column=2)
        cell.value = role
        cell.font = Font(size=9)
        cell.border = thin_border

        # C列: 雇用形態
        cell = ws.cell(row=row, column=3)
        cell.value = emp
        cell.font = Font(size=9)
        cell.border = thin_border

        # D列: 氏名
        cell = ws.cell(row=row, column=4)
        cell.value = name
        cell.font = Font(size=10, bold=True)
        cell.border = thin_border

        # 背景色
        if role == '調理職員':
            bg_color = COLORS['cook_bg']
        elif role in ['管理者', '計画作成担当者', '機能訓練指導員',
                       '生活相談員', '看護職員', '事務職員']:
            bg_color = COLORS['admin_bg']
        else:
            bg_color = None

        # シフトカウント用
        shift_counts = defaultdict(int)

        # 日別シフト
        for day in range(1, num_days + 1):
            col = DAY_COL_START + day - 1
            assigned = schedule.get(day, '')
            display = get_shift_display(assigned) if assigned else ''

            cell = ws.cell(row=row, column=col)
            cell.value = display
            if assigned:
                cell.font = get_cell_font(assigned)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            if is_weekend(year, month, day):
                cell.fill = PatternFill(
                    start_color=COLORS['weekend_bg'],
                    end_color=COLORS['weekend_bg'], fill_type='solid')
            elif bg_color:
                cell.fill = PatternFill(
                    start_color=bg_color,
                    end_color=bg_color, fill_type='solid')

            # カウント
            if assigned:
                shift_counts[assigned] += 1

        # 集計列: Ａ, Ｂ, Ｄ, Ｅ, C, 夜, 明, 他, 調1, 調2, 有, 休, 出勤計, 出勤上限, チェック
        count_keys = ['A', 'B', 'D', 'E', 'C', 'night', 'dawn',
                      None, 'cook1', 'cook2', '有', None, None, None, None]

        # 「他」= 研 + 健
        other_count = shift_counts.get('研', 0) + shift_counts.get('健', 0)
        # 「休」= 休 + 希
        off_count = shift_counts.get('休', 0) + shift_counts.get('希', 0)
        # 出勤計
        work_total = summary['work_days']

        values = [
            shift_counts.get('A', 0),
            shift_counts.get('B', 0),
            shift_counts.get('D', 0),
            shift_counts.get('E', 0),
            shift_counts.get('C', 0),
            shift_counts.get('night', 0),
            shift_counts.get('dawn', 0),
            other_count,
            shift_counts.get('cook1', 0),
            shift_counts.get('cook2', 0),
            shift_counts.get('有', 0),
            off_count,
            work_total,
            '',  # 出勤上限（設定ファイルから取得可能）
            '',  # 出勤日数チェック
        ]

        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=summary_col_start + i)
            cell.value = val
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    current_row = DATA_ROW_START + len(staff_rows) + 1

    # ============================================================
    # 介護職員条件 集計ブロック
    # ============================================================
    ws.cell(row=current_row, column=2).value = "介護職員条件満たしているか"
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)

    care_shift_labels = [
        ('A', 'A(7:30-15:30)'),
        ('B', 'B(9:30-17:30)'),
        ('C', 'C(12:30-20:30)'),
        ('D', 'D(8:30-16:30)'),
        ('night', '夜勤(17:00-1:00)'),
        ('dawn', '明け(1:00-9:00)'),
    ]

    for si, (shift, label) in enumerate(care_shift_labels):
        row = current_row + si
        ws.cell(row=row, column=3).value = label
        ws.cell(row=row, column=3).font = Font(size=9)
        ws.cell(row=row, column=3).border = thin_border

        for day in range(1, num_days + 1):
            col = DAY_COL_START + day - 1
            count = 0
            for sr in result['staff']:
                if sr['role'] == '介護職員':
                    if sr['schedule'].get(day) == shift:
                        count += 1
            cell = ws.cell(row=row, column=col)
            cell.value = count
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.fill = PatternFill(
                start_color=COLORS['summary_bg'],
                end_color=COLORS['summary_bg'], fill_type='solid')

    current_row += len(care_shift_labels)

    # 出勤合計行
    ws.cell(row=current_row, column=3).value = "出勤合計"
    ws.cell(row=current_row, column=3).font = Font(bold=True, size=9)
    ws.cell(row=current_row, column=3).border = thin_border
    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        count = sum(1 for sr in result['staff']
                    if sr['schedule'].get(day) in ALL_SHIFTS + WORK_COUNTED)
        cell = ws.cell(row=current_row, column=col)
        cell.value = count
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = PatternFill(
            start_color=COLORS['summary_bg'],
            end_color=COLORS['summary_bg'], fill_type='solid')
    current_row += 1

    # 休日・有給行
    ws.cell(row=current_row, column=3).value = "休日・有給"
    ws.cell(row=current_row, column=3).font = Font(size=9)
    ws.cell(row=current_row, column=3).border = thin_border
    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        count = sum(1 for sr in result['staff']
                    if sr['schedule'].get(day) in OFF_TYPES + ['有'])
        cell = ws.cell(row=current_row, column=col)
        cell.value = count
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    current_row += 1

    # シフト成立行
    ws.cell(row=current_row, column=3).value = "シフト成立"
    ws.cell(row=current_row, column=3).font = Font(bold=True, size=9)
    ws.cell(row=current_row, column=3).border = thin_border
    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        is_sun = is_sunday(year, month, day)

        # 介護職員のみでシフト判定
        cc = defaultdict(int)
        for sr in result['staff']:
            if sr['role'] == '介護職員':
                a = sr['schedule'].get(day, '休')
                if a in CARE_SHIFTS:
                    cc[a] += 1

        all_filled = all(cc[s] > 0 for s in CARE_SHIFTS
                         if not (is_sun and s == 'A'))
        best = ((cc['B'] > 0 and cc['C'] > 0 and
                 cc['night'] > 0 and cc['dawn'] > 0) or
                (cc['C'] > 0 and cc['D'] > 0 and
                 cc['night'] > 0 and cc['dawn'] > 0))

        cell = ws.cell(row=current_row, column=col)
        if not all_filled:
            cell.value = "不成立"
            cell.fill = PatternFill(
                start_color=COLORS['violation_bg'],
                end_color=COLORS['violation_bg'], fill_type='solid')
        elif best:
            cell.value = "成立"
            cell.fill = PatternFill(
                start_color=COLORS['success_bg'],
                end_color=COLORS['success_bg'], fill_type='solid')
        else:
            cell.value = "30分残業"
            cell.fill = PatternFill(
                start_color=COLORS['overtime_bg'],
                end_color=COLORS['overtime_bg'], fill_type='solid')
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    current_row += 2

    # ============================================================
    # 介護関連職員出勤数ブロック
    # ============================================================
    ws.cell(row=current_row, column=2).value = "介護関連職員出勤数"
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)

    all_shift_labels = [
        ('A', 'A(7:30-15:30)'),
        ('B', 'B(9:30-17:30)'),
        ('C', 'C(12:30-20:30)'),
        ('D', 'D(8:30-16:30)'),
        ('E', 'E(8:30-16:30)'),
        ('night', '夜勤(17:00-1:00)'),
        ('dawn', '明け(1:00-9:00)'),
    ]

    care_related_roles = ['介護職員', '管理者', '計画作成担当者',
                          '機能訓練指導員', '生活相談員', '看護職員']

    for si, (shift, label) in enumerate(all_shift_labels):
        row = current_row + si
        ws.cell(row=row, column=3).value = label
        ws.cell(row=row, column=3).font = Font(size=9)
        ws.cell(row=row, column=3).border = thin_border

        for day in range(1, num_days + 1):
            col = DAY_COL_START + day - 1
            count = 0
            for sr in result['staff']:
                if sr['role'] in care_related_roles:
                    if sr['schedule'].get(day) == shift:
                        count += 1
            cell = ws.cell(row=row, column=col)
            cell.value = count
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    current_row += len(all_shift_labels)

    # 介護関連出勤合計
    ws.cell(row=current_row, column=3).value = "出勤合計"
    ws.cell(row=current_row, column=3).font = Font(bold=True, size=9)
    ws.cell(row=current_row, column=3).border = thin_border
    for day in range(1, num_days + 1):
        col = DAY_COL_START + day - 1
        count = 0
        for sr in result['staff']:
            if sr['role'] in care_related_roles:
                a = sr['schedule'].get(day, '休')
                if a in ALL_SHIFTS or a in WORK_COUNTED:
                    count += 1
        cell = ws.cell(row=current_row, column=col)
        cell.value = count
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    current_row += 2

    # ============================================================
    # 不備報告
    # ============================================================
    ws.cell(row=current_row, column=2).value = "【不備報告】"
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=11,
                                                    color='FF0000')
    current_row += 1
    violations = result.get('violations', [])
    if violations:
        for v in violations:
            ws.cell(row=current_row, column=2).value = f"⚠ {v}"
            ws.cell(row=current_row, column=2).font = Font(
                size=10, color='FF0000')
            current_row += 1
    else:
        ws.cell(row=current_row, column=2).value = "✅ ルール違反なし"
        ws.cell(row=current_row, column=2).font = Font(
            size=10, color='009900')

    # ============================================================
    # リスト表シート
    # ============================================================
    ws_list = wb.create_sheet("リスト表")
    list_data = [
        ['Ａ', 'Ｅ', '調1', '', '常勤/専従'],
        ['Ｂ', '希', '調2', '', '常勤/兼任'],
        ['Ｃ', '休', '', '', '非常勤/専従'],
        ['Ｄ', '有', '', '', '非常勤/兼任'],
        ['Ｅ', '', '', '', ''],
        ['夜', '', '', '', ''],
        ['明', '', '', '', ''],
        ['希', '', '', '', ''],
        ['休', '', '', '', ''],
        ['有', '', '', '', ''],
        ['健', '', '', '', ''],
        ['研', '', '', '', ''],
        ['調1', '', '', '', ''],
        ['調2', '', '', '', ''],
    ]
    for r, row_data_list in enumerate(list_data, 2):
        for c, val in enumerate(row_data_list, 1):
            ws_list.cell(row=r, column=c).value = val

    # ============================================================
    # 保存
    # ============================================================
    wb.save(output_path)
    print(f"シフト表を出力しました: {output_path}")
