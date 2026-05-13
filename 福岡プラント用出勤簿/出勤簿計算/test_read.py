import openpyxl

wb = openpyxl.load_workbook('../【出勤簿】202604_計算式適用済.xlsx', data_only=True)
ws = wb['坂井　謙斗']
for r in range(21, 29):
    row_data = {
        'D': ws[f'D{r}'].value,
        'BD': ws[f'BD{r}'].value,
        'BE': ws[f'BE{r}'].value,
        'BG': ws[f'BG{r}'].value,
        'BH': ws[f'BH{r}'].value,
        'BI': ws[f'BI{r}'].value,
        'R': ws[f'R{r}'].value,
        'S': ws[f'S{r}'].value,
        'T': ws[f'T{r}'].value,
        'V': ws[f'V{r}'].value
    }
    print(f"Row {r}: D={row_data['D']}, BD={row_data['BD']}, BE={row_data['BE']}, BG={row_data['BG']}, BH={row_data['BH']}, BI={row_data['BI']}, R={row_data['R']}, S={row_data['S']}, T={row_data['T']}, V={row_data['V']}")
