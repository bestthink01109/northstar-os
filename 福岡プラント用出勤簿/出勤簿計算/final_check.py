import openpyxl

path = "/Users/fuminariaksse/Desktop/antigravity Folder/福岡プラント用出勤簿/出勤簿計算/【出勤簿】202605_計算式適用済.xlsx"
try:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["実績"]
    # 土本さんの行（3行目〜）をチェック
    print("--- 転記結果の確認 (5月8日付近) ---")
    # C列が1日なので、J列が8日
    print(f"氏名: {ws['A3'].value}")
    print(f"5/8 出勤状態: {ws['J3'].value}")
    print(f"5/8 現場名: {ws['J4'].value}")
    print(f"5/8 残業: {ws['J5'].value}")
except Exception as e:
    print(f"Error: {e}")
