import openpyxl

path = "/Users/fuminariaksse/Desktop/antigravity Folder/福岡プラント用出勤簿/出勤簿計算/【出勤簿】202605.xlsx"
try:
    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    if "実績" in wb.sheetnames:
        ws = wb["実績"]
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            print(f"Row: {row}")
    else:
        print("Error: '実績' sheet not found.")
except Exception as e:
    print(f"Error: {e}")
