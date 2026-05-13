import openpyxl
wb = openpyxl.load_workbook('../【出勤簿】202604_計算式適用済.xlsx', data_only=True)
ws = wb['坂井　謙斗']
print("Row | Date | Status | R (所定) | S (法内) | T (残業) | U (深夜) | V (休日) | X (不就労)")
for r in range(4, 25):
    b = ws[f'B{r}'].value
    d = ws[f'D{r}'].value
    if b is None: continue
    
    r_val = ws[f'R{r}'].value
    s_val = ws[f'S{r}'].value
    t_val = ws[f'T{r}'].value
    u_val = ws[f'U{r}'].value
    v_val = ws[f'V{r}'].value
    x_val = ws[f'X{r}'].value
    print(f"{r} | {b.strftime('%m/%d')} | {d} | R={r_val} | S={s_val} | T={t_val} | U={u_val} | V={v_val} | X={x_val}")
